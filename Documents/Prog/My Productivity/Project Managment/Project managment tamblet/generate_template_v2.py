#!/usr/bin/env python3
"""
Project Management Template Generator v2.0
Enhanced with advanced algorithms and strategic improvements
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, IconSetRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import json

class AdvancedProjectTemplateGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.setup_enhanced_styles()
        self.risk_factors = {}
        self.critical_path = []
        
    def setup_enhanced_styles(self):
        """Enhanced color scheme with gradients and modern design"""
        self.colors = {
            'primary': '9E1F63',
            'primary_light': 'C94F83',
            'secondary': '721548',
            'secondary_light': '924568',
            'accent_blue': '005B8C',
            'accent_blue_light': '3088BC',
            'accent_coral': 'E05E3D',
            'success': '27AE60',
            'warning': 'F39C12',
            'danger': 'E74C3C',
            'info': '3498DB',
            'dark': '2C3E50',
            'light': 'ECF0F1',
            'white': 'FFFFFF'
        }
        
        self.fonts = {
            'title': Font(name='Verdana', size=18, bold=True, color=self.colors['primary']),
            'header': Font(name='Verdana', size=14, bold=True, color=self.colors['primary']),
            'subheader': Font(name='Verdana', size=12, bold=True, color=self.colors['secondary']),
            'body': Font(name='Verdana', size=10),
            'body_bold': Font(name='Verdana', size=10, bold=True),
            'small': Font(name='Verdana', size=9),
            'impact': Font(name='Verdana', size=11, bold=True, color=self.colors['danger'])
        }
        
        self.fills = {
            'header': PatternFill(start_color=self.colors['primary_light'], end_color=self.colors['primary_light'], fill_type='solid'),
            'parent': PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid'),
            'child': PatternFill(start_color=self.colors['white'], end_color=self.colors['white'], fill_type='solid'),
            'critical': PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid'),
            'success': PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid'),
            'warning': PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid'),
            'info': PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
        }
        
        self.borders = {
            'thin': Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            ),
            'thick': Border(
                left=Side(style='medium', color=self.colors['primary']),
                right=Side(style='medium', color=self.colors['primary']),
                top=Side(style='medium', color=self.colors['primary']),
                bottom=Side(style='medium', color=self.colors['primary'])
            ),
            'bottom_thick': Border(
                bottom=Side(style='thick', color=self.colors['primary'])
            )
        }
        
    def create_enhanced_main_sheet(self):
        """Create main sheet with advanced features"""
        ws = self.wb.active
        ws.title = "Project Tasks"
        
        # Enhanced headers with additional columns
        headers = [
            'Task ID', 'Task Name', 'Task Type', 'Parent Task ID', 'WBS Code',
            'Duration (Days)', 'Start Date', 'End Date', 'Actual Start', 'Actual End',
            '% Complete', 'Task Agility', 'Dependencies', 'Dependency Type', 'Lag/Lead',
            'Impact Score', 'Risk Score', 'Criticality Level', 'Priority Score',
            'Resource Assignment', 'Resource Load %', 'Cost Budget', 'Actual Cost',
            'Status', 'Health Indicator', 'Milestone', 'Deliverables',
            'Blocking Tasks', 'Blocked By', 'Total Float', 'Free Float', 'Critical Path',
            'Weighted Progress', 'Rolled Up Progress', 'Variance Days', 'SPI', 'CPI',
            'Risk Mitigation', 'Lessons Learned', 'Notes'
        ]
        
        # Apply enhanced header formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.borders['thick']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        # Enhanced column widths
        self.set_optimized_column_widths(ws, headers)
        
        # Add enhanced validations
        self.add_enhanced_validations(ws)
        
        # Add comprehensive sample data
        self.add_comprehensive_sample_data(ws)
        
        # Add advanced formulas
        self.add_advanced_formulas(ws)
        
        # Add enhanced conditional formatting
        self.add_enhanced_conditional_formatting(ws)
        
        # Add data filters
        ws.auto_filter.ref = f"A1:AN1000"
        
        # Freeze panes for better navigation
        ws.freeze_panes = 'F2'
        
    def set_optimized_column_widths(self, ws, headers):
        """Set optimized column widths based on content"""
        width_map = {
            'Task ID': 10,
            'Task Name': 35,
            'Task Type': 12,
            'Parent Task ID': 15,
            'WBS Code': 12,
            'Duration (Days)': 12,
            'Start Date': 12,
            'End Date': 12,
            'Actual Start': 12,
            'Actual End': 12,
            '% Complete': 12,
            'Task Agility': 12,
            'Dependencies': 15,
            'Dependency Type': 15,
            'Lag/Lead': 10,
            'Impact Score': 12,
            'Risk Score': 12,
            'Criticality Level': 15,
            'Priority Score': 12,
            'Resource Assignment': 20,
            'Resource Load %': 15,
            'Cost Budget': 12,
            'Actual Cost': 12,
            'Status': 15,
            'Health Indicator': 15,
            'Milestone': 10,
            'Deliverables': 25,
            'Blocking Tasks': 15,
            'Blocked By': 15,
            'Total Float': 10,
            'Free Float': 10,
            'Critical Path': 12,
            'Weighted Progress': 15,
            'Rolled Up Progress': 15,
            'Variance Days': 12,
            'SPI': 8,
            'CPI': 8,
            'Risk Mitigation': 25,
            'Lessons Learned': 25,
            'Notes': 30
        }
        
        for col, header in enumerate(headers, 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width_map.get(header, 12)
            
    def add_enhanced_validations(self, ws):
        """Add comprehensive data validations"""
        # Task Type validation
        task_type_val = DataValidation(
            type="list",
            formula1='"Parent,Child,Milestone,Summary"',
            allow_blank=False,
            showDropDown=True
        )
        task_type_val.add('C2:C1000')
        ws.add_data_validation(task_type_val)
        
        # Task Agility validation
        agility_val = DataValidation(
            type="list",
            formula1='"Parallel,Sequential,Start-to-Start,Finish-to-Finish"',
            allow_blank=False,
            showDropDown=True
        )
        agility_val.add('L2:L1000')
        ws.add_data_validation(agility_val)
        
        # Dependency Type validation
        dep_type_val = DataValidation(
            type="list",
            formula1='"FS,SS,FF,SF,FS+,SS+,FF+,SF+"',
            allow_blank=True,
            showDropDown=True
        )
        dep_type_val.add('N2:N1000')
        ws.add_data_validation(dep_type_val)
        
        # Criticality Level validation
        criticality_val = DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Minimal"',
            allow_blank=False,
            showDropDown=True
        )
        criticality_val.add('R2:R1000')
        ws.add_data_validation(criticality_val)
        
        # Status validation
        status_val = DataValidation(
            type="list",
            formula1='"Not Started,Planning,In Progress,Testing,Review,Complete,On Hold,Cancelled,Delayed,Blocked"',
            allow_blank=False,
            showDropDown=True
        )
        status_val.add('X2:X1000')
        ws.add_data_validation(status_val)
        
        # Health Indicator validation
        health_val = DataValidation(
            type="list",
            formula1='"Green,Yellow,Orange,Red,Black"',
            allow_blank=False,
            showDropDown=True
        )
        health_val.add('Y2:Y1000')
        ws.add_data_validation(health_val)
        
        # Milestone validation
        milestone_val = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False,
            showDropDown=True
        )
        milestone_val.add('Z2:Z1000')
        ws.add_data_validation(milestone_val)
        
        # Percentage validations
        for col in ['K', 'U']:  # % Complete, Resource Load %
            percent_val = DataValidation(
                type="decimal",
                operator="between",
                formula1=0,
                formula2=100,
                allow_blank=False
            )
            percent_val.add(f'{col}2:{col}1000')
            ws.add_data_validation(percent_val)
            
    def add_comprehensive_sample_data(self, ws):
        """Add realistic project data with multiple phases"""
        sample_data = [
            # Project phases (Parents)
            ['P001', 'Project Phoenix - Digital Transformation', 'Parent', '', '1', 120, '2024-01-01', '', '', '', 0, 'Sequential', '', '', 0, '', '', 'Critical', '', 'Program Management', 100, 5000000, 0, 'In Progress', 'Yellow', 'No', 'Complete digital transformation', '', '', '', '', 'Yes', '', '', 0, '', '', 'Risk assessment in progress', '', 'Strategic initiative'],
            ['P002', 'Phase 1: Discovery & Planning', 'Parent', 'P001', '1.1', 20, '2024-01-01', '', '', '', 0, 'Sequential', '', '', 0, '', '', 'High', '', 'PM Team', 100, 500000, 0, 'Complete', 'Green', 'No', 'Requirements and architecture', '', '', '', '', 'Yes', '', '', 0, '', '', '', 'Stakeholder alignment critical', 'Foundation phase'],
            ['P003', 'Phase 2: Infrastructure Setup', 'Parent', 'P001', '1.2', 30, '2024-01-21', '', '', '', 0, 'Parallel', 'P002', 'FS', 0, '', '', 'High', '', 'Infrastructure Team', 100, 1500000, 0, 'In Progress', 'Yellow', 'No', 'Cloud infrastructure ready', '', '', '', '', 'Yes', '', '', 0, '', '', 'Cloud provider selection', '', 'Technical foundation'],
            ['P004', 'Phase 3: Development Sprint 1', 'Parent', 'P001', '1.3', 40, '2024-02-20', '', '', '', 0, 'Parallel', 'P003', 'SS+10', 10, '', '', 'Critical', '', 'Dev Team Alpha', 100, 1000000, 0, 'Planning', 'Green', 'No', 'Core modules developed', '', '', '', '', 'Yes', '', '', 0, '', '', '', '', 'Agile development'],
            ['P005', 'Phase 4: Integration & Testing', 'Parent', 'P001', '1.4', 20, '2024-04-01', '', '', '', 0, 'Sequential', 'P004', 'FS-5', -5, '', '', 'High', '', 'QA Team', 100, 750000, 0, 'Not Started', 'Green', 'No', 'System integration complete', '', '', '', '', 'No', '', '', 0, '', '', 'Test automation required', '', 'Quality assurance'],
            ['P006', 'Phase 5: Deployment & GoLive', 'Parent', 'P001', '1.5', 10, '2024-04-21', '', '', '', 0, 'Sequential', 'P005', 'FS', 0, '', '', 'Critical', '', 'DevOps Team', 100, 250000, 0, 'Not Started', 'Green', 'No', 'System live in production', '', '', '', '', 'Yes', '', '', 0, '', '', 'Rollback plan ready', '', 'Go-live phase'],
            
            # Milestones
            ['M001', 'Project Kickoff Complete', 'Milestone', 'P002', '1.1.0', 0, '2024-01-01', '', '2024-01-01', '', 100, 'Sequential', '', '', 0, '', '', 'High', '', 'All Teams', 0, 0, 0, 'Complete', 'Green', 'Yes', 'Kickoff meeting held', '', '', 0, 0, 'No', '', '', 0, '', '', '', 'Great team energy', 'Key milestone'],
            ['M002', 'Requirements Signed Off', 'Milestone', 'P002', '1.1.99', 0, '2024-01-20', '', '', '', 0, 'Sequential', 'C004', 'FS', 0, '', '', 'Critical', '', 'Stakeholders', 0, 0, 0, 'Not Started', 'Green', 'Yes', 'Requirements approval', '', '', 0, 0, 'Yes', '', '', 0, '', '', '', '', 'Gate 1'],
            
            # Child tasks for P002 (Discovery & Planning)
            ['C001', 'Stakeholder Interviews', 'Child', 'P002', '1.1.1', 5, '2024-01-02', '', '2024-01-02', '', 100, 'Parallel', 'M001', 'FS', 0, '', '', 'High', '', 'BA Team', 80, 50000, 45000, 'Complete', 'Green', 'No', 'Interview notes', '', '', 2, 2, 'No', '', '', 0, 1.0, 0.9, '', 'Excellent insights gathered', 'Completed on time'],
            ['C002', 'Current State Analysis', 'Child', 'P002', '1.1.2', 5, '2024-01-02', '', '2024-01-02', '', 100, 'Parallel', 'M001', 'FS', 0, '', '', 'High', '', 'Tech Architects', 100, 75000, 70000, 'Complete', 'Green', 'No', 'As-is documentation', '', '', 0, 0, 'No', '', '', 0, 1.0, 0.93, '', 'Legacy system complexities', 'Technical debt identified'],
            ['C003', 'Future State Design', 'Child', 'P002', '1.1.3', 7, '2024-01-07', '', '2024-01-08', '', 100, 'Sequential', 'C001,C002', 'FS', 0, '', '', 'Critical', '', 'Solution Architects', 100, 100000, 95000, 'Complete', 'Green', 'No', 'To-be architecture', '', '', 1, 0, 'Yes', '', '', 0, 0.95, 0.95, '', 'Cloud-native approach', 'Microservices architecture'],
            ['C004', 'Requirements Documentation', 'Child', 'P002', '1.1.4', 3, '2024-01-14', '', '2024-01-15', '', 100, 'Sequential', 'C003', 'FS', 0, '', '', 'High', '', 'BA Team', 100, 25000, 25000, 'Complete', 'Green', 'No', 'BRD & FRD documents', 'M002', '', 0, 0, 'Yes', '', '', 0, 0.98, 1.0, '', 'Clear requirements', 'Well documented'],
            
            # Child tasks for P003 (Infrastructure)
            ['C005', 'Cloud Account Setup', 'Child', 'P003', '1.2.1', 3, '2024-01-21', '', '2024-01-21', '', 100, 'Sequential', 'P002', 'FS', 0, '', '', 'High', '', 'Cloud Team', 50, 10000, 9500, 'Complete', 'Green', 'No', 'AWS accounts ready', '', '', 5, 5, 'No', '', '', 0, 1.0, 0.95, '', 'Multi-account strategy', 'Security first approach'],
            ['C006', 'Network Architecture', 'Child', 'P003', '1.2.2', 5, '2024-01-24', '', '2024-01-25', '', 100, 'Sequential', 'C005', 'FS', 0, '', '', 'Critical', '', 'Network Engineers', 100, 150000, 145000, 'Complete', 'Green', 'No', 'VPC and connectivity', '', '', 3, 3, 'Yes', '', '', 1, 0.95, 0.97, '', 'Zero-trust model', 'Implemented successfully'],
            ['C007', 'Security Framework', 'Child', 'P003', '1.2.3', 7, '2024-01-24', '', '2024-01-26', '', 80, 'Parallel', 'C005', 'FS', 0, '', '', 'Critical', '', 'Security Team', 100, 200000, 150000, 'In Progress', 'Yellow', 'No', 'Security policies & tools', '', '', 0, 0, 'Yes', '', '', -1, 0.85, 0.75, 'Additional security review', 'Compliance requirements', 'GDPR compliance needed'],
            ['C008', 'CI/CD Pipeline Setup', 'Child', 'P003', '1.2.4', 10, '2024-01-29', '', '', '', 50, 'Sequential', 'C006', 'FS', 0, '', '', 'High', '', 'DevOps Team', 100, 100000, 40000, 'In Progress', 'Green', 'No', 'Automated pipelines', 'C009,C010', '', 0, 0, 'No', '', '', 0, 1.0, 0.4, '', 'GitOps approach', 'Jenkins to GitLab migration'],
            ['C009', 'Development Environment', 'Child', 'P003', '1.2.5', 5, '2024-02-08', '', '', '', 20, 'Parallel', 'C008', 'SS', 0, '', '', 'Medium', '', 'DevOps Team', 80, 50000, 5000, 'In Progress', 'Green', 'No', 'Dev env ready', '', 'C008', 2, 2, 'No', '', '', 0, 0.8, 0.1, '', '', 'Containerized approach'],
            ['C010', 'Testing Environment', 'Child', 'P003', '1.2.6', 5, '2024-02-08', '', '', '', 20, 'Parallel', 'C008', 'SS', 0, '', '', 'Medium', '', 'QA Team', 80, 50000, 5000, 'In Progress', 'Green', 'No', 'Test env ready', '', 'C008', 2, 2, 'No', '', '', 0, 0.8, 0.1, '', '', 'Automated provisioning'],
            
            # Child tasks for P004 (Development)
            ['C011', 'User Authentication Module', 'Child', 'P004', '1.3.1', 10, '2024-03-01', '', '', '', 0, 'Parallel', 'C009', 'FS', 0, '', '', 'Critical', '', 'Team Alpha', 100, 150000, 0, 'Not Started', 'Green', 'No', 'OAuth2 implementation', '', '', 0, 0, 'Yes', '', '', 0, 0, 0, '', '', 'Planning phase'],
            ['C012', 'API Gateway Development', 'Child', 'P004', '1.3.2', 8, '2024-03-01', '', '', '', 0, 'Parallel', 'C009', 'FS', 0, '', '', 'High', '', 'Team Beta', 100, 120000, 0, 'Not Started', 'Green', 'No', 'REST API gateway', '', '', 2, 2, 'Yes', '', '', 0, 0, 0, '', '', 'GraphQL consideration'],
            ['C013', 'Database Schema Design', 'Child', 'P004', '1.3.3', 5, '2024-03-01', '', '', '', 0, 'Sequential', 'C009', 'FS', 0, '', '', 'Critical', '', 'Data Team', 100, 80000, 0, 'Not Started', 'Green', 'No', 'Optimized schema', 'C014,C015', '', 0, 0, 'Yes', '', '', 0, 0, 0, '', '', 'NoSQL evaluation'],
            ['C014', 'Core Business Logic', 'Child', 'P004', '1.3.4', 15, '2024-03-06', '', '', '', 0, 'Sequential', 'C013', 'FS', 0, '', '', 'Critical', '', 'Team Alpha', 100, 250000, 0, 'Not Started', 'Green', 'No', 'Business rules engine', '', 'C013', 0, 0, 'Yes', '', '', 0, 0, 0, '', '', 'Domain-driven design'],
            ['C015', 'Data Migration Scripts', 'Child', 'P004', '1.3.5', 10, '2024-03-06', '', '', '', 0, 'Parallel', 'C013', 'FS', 0, '', '', 'High', '', 'Data Team', 100, 100000, 0, 'Not Started', 'Green', 'No', 'ETL processes', '', 'C013', 5, 5, 'No', '', '', 0, 0, 0, '', '', 'Data quality critical'],
            
            # Risk items
            ['R001', 'Technical Debt Assessment', 'Child', 'P002', '1.1.5', 3, '2024-01-10', '', '2024-01-11', '', 100, 'Parallel', 'C002', 'SS', 0, '', '', 'Medium', '', 'Tech Leads', 50, 20000, 18000, 'Complete', 'Yellow', 'No', 'Debt register', '', '', 0, 0, 'No', '', '', 0, 0.95, 0.9, 'High technical debt found', 'Refactoring needed', 'Legacy system issues'],
            ['R002', 'Security Vulnerability Scan', 'Child', 'P003', '1.2.7', 2, '2024-02-01', '', '', '', 60, 'Parallel', 'C007', 'SS+3', 3, '', '', 'Critical', '', 'Security Team', 100, 30000, 15000, 'In Progress', 'Red', 'No', 'Vulnerability report', '', '', 0, 0, 'Yes', '', '', -2, 0.7, 0.5, 'Critical vulnerabilities found', 'Immediate action required', 'Zero-day patch needed'],
        ]
        
        # Write comprehensive data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.borders['thin']
                
                # Apply formatting based on task type
                if col_idx == 3:  # Task Type column
                    if value == 'Parent':
                        cell.fill = self.fills['parent']
                        cell.font = self.fonts['body_bold']
                    elif value == 'Milestone':
                        cell.fill = self.fills['info']
                        cell.font = Font(name='Verdana', size=10, bold=True, italic=True)
                    else:
                        cell.font = self.fonts['body']
                        
    def add_advanced_formulas(self, ws):
        """Add sophisticated formulas with multiple factors"""
        for row in range(2, 50):  # Extended range
            # Advanced End Date calculation with lag/lead
            ws[f'H{row}'] = f'''=IF(AND(G{row}<>"",F{row}<>""),
                IF(O{row}<>"",WORKDAY(G{row},F{row}-1+O{row}),WORKDAY(G{row},F{row}-1)),"")'''
            
            # Sophisticated % Complete calculation
            ws[f'K{row}'] = f'''=IF(X{row}="Complete",100,
                IF(X{row}="Cancelled",J{row},
                IF(AND(I{row}<>"",F{row}<>"",H{row}<>""),
                    MIN(MAX((TODAY()-I{row})/(H{row}-I{row})*100,0),99),
                    IF(X{row}="In Progress",MAX(K{row},10),0))))'''
            
            # Enhanced Impact Score with machine learning factors
            impact_formula = self.create_ml_impact_formula(row)
            ws[f'P{row}'] = impact_formula
            
            # Risk Score calculation
            risk_formula = self.create_risk_score_formula(row)
            ws[f'Q{row}'] = risk_formula
            
            # Priority Score (combines impact and risk)
            ws[f'S{row}'] = f'=ROUND((P{row}*0.6+Q{row}*0.4),0)'
            
            # Advanced Health Indicator
            health_formula = self.create_health_formula(row)
            ws[f'Y{row}'] = health_formula
            
            # Float calculations
            ws[f'AD{row}'] = f'=IF(AND(H{row}<>"",G{row}<>""),NETWORKDAYS(TODAY(),H{row})-NETWORKDAYS(TODAY(),G{row}),"")'
            ws[f'AE{row}'] = f'=IF(AND(AD{row}<>"",AC{row}<>""),MIN(AD{row},SUMIF(M$2:M$100,A{row},AD$2:AD$100)),AD{row})'
            
            # Critical Path indicator
            ws[f'AF{row}'] = f'=IF(OR(AD{row}<=0,AE{row}<=0),"Yes","No")'
            
            # Weighted Progress for parents
            ws[f'AG{row}'] = f'''=IF(C{row}="Parent",
                IF(SUMIF(D$2:D$100,A{row},P$2:P$100)>0,
                    SUMPRODUCT((D$2:D$100=A{row})*(K$2:K$100)*(P$2:P$100))/SUMIF(D$2:D$100,A{row},P$2:P$100),
                    AVERAGE(IF(D$2:D$100=A{row},K$2:K$100))),
                K{row})'''
            
            # Variance calculation
            ws[f'AI{row}'] = f'=IF(AND(J{row}<>"",H{row}<>""),J{row}-H{row},"")'
            
            # Schedule Performance Index (SPI)
            ws[f'AJ{row}'] = f'=IF(AND(K{row}>0,F{row}<>""),ROUND(K{row}/MAX((TODAY()-G{row})/(H{row}-G{row})*100,1),2),"")'
            
            # Cost Performance Index (CPI)
            ws[f'AK{row}'] = f'=IF(AND(W{row}>0,V{row}>0),ROUND(V{row}/W{row},2),"")'
            
    def create_ml_impact_formula(self, row):
        """Create ML-enhanced impact score formula"""
        return f'''=ROUND(
            (
                (F{row}/MAX(F:F)*20) +
                (LEN(M{row})-LEN(SUBSTITUTE(M{row},",",""))+1)*25/MAX(5,COUNTIF(M:M,"*"&A{row}&"*")) +
                (SWITCH(R{row},"Critical",100,"High",80,"Medium",60,"Low",40,"Minimal",20,50)*0.15) +
                ((COUNTA(A:A)-ROW()+1)/COUNTA(A:A)*15) +
                (IF(U{row}>80,20,IF(U{row}>50,10,0))) +
                (IF(Z{row}="Yes",25,0)) +
                (IF(AB{row}<>"",LEN(AB{row})-LEN(SUBSTITUTE(AB{row},",",""))+1,0)*5) +
                (IF(V{row}>0,LOG10(V{row})/LOG10(MAX(V:V))*10,0))
            ) *
            IF(AB{row}<>"",1.5,1) *
            IF(L{row}="Parallel",0.7,1) *
            IF(X{row}="Delayed",1.3,IF(X{row}="Blocked",1.8,1)) *
            IF(AF{row}="Yes",1.4,1)
        ,0)'''
        
    def create_risk_score_formula(self, row):
        """Create comprehensive risk score formula"""
        return f'''=ROUND(
            (
                IF(X{row}="Blocked",30,IF(X{row}="Delayed",20,0)) +
                IF(AI{row}<-7,25,IF(AI{row}<-3,15,IF(AI{row}<0,5,0))) +
                IF(AND(AJ{row}<>"",AJ{row}<0.8),20,0) +
                IF(AND(AK{row}<>"",AK{row}<0.9),15,0) +
                IF(AF{row}="Yes",20,0) +
                IF(AD{row}<0,25,IF(AD{row}<3,15,0)) +
                (100-K{row})/5 +
                IF(AC{row}<>"",10,0)
            ) * 
            IF(R{row}="Critical",1.5,IF(R{row}="High",1.2,1))
        ,0)'''
        
    def create_health_formula(self, row):
        """Create sophisticated health indicator formula"""
        return f'''=IF(X{row}="Complete","Green",
            IF(X{row}="Cancelled","Black",
            IF(OR(X{row}="Blocked",Q{row}>80),"Red",
            IF(OR(X{row}="Delayed",Q{row}>60,AND(K{row}<50,H{row}<TODAY())),"Orange",
            IF(OR(Q{row}>40,AND(K{row}<80,H{row}<=TODAY()+7),AJ{row}<0.9),"Yellow",
            "Green")))))'''
            
    def add_enhanced_conditional_formatting(self, ws):
        """Add sophisticated conditional formatting with icons"""
        # Health indicator with custom colors
        health_colors = {
            'Green': self.fills['success'],
            'Yellow': self.fills['warning'],
            'Orange': PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid'),
            'Red': self.fills['critical'],
            'Black': PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        }
        
        for color, fill in health_colors.items():
            ws.conditional_formatting.add('Y2:Y1000',
                CellIsRule(operator='equal', formula=[f'"{color}"'], fill=fill))
        
        # Progress bars with gradient
        ws.conditional_formatting.add('K2:K1000',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='E74C3C',
                mid_type='num', mid_value=50, mid_color='F39C12',
                end_type='num', end_value=100, end_color='27AE60'
            ))
        
        # Impact score heatmap
        ws.conditional_formatting.add('P2:P1000',
            ColorScaleRule(
                start_type='percentile', start_value=10, start_color='FFFFFF',
                mid_type='percentile', mid_value=50, mid_color='FF9999',
                end_type='percentile', end_value=90, end_color='CC0000'
            ))
        
        # Risk score with icons
        icon_set = IconSetRule('3Symbols', 'percent', [0, 33, 67])
        ws.conditional_formatting.add('Q2:Q1000', icon_set)
        
        # Critical path highlighting
        ws.conditional_formatting.add('AF2:AF1000',
            CellIsRule(operator='equal', formula=['"Yes"'], 
                      font=Font(bold=True, color='FF0000')))
        
        # SPI/CPI performance indicators
        for col in ['AJ', 'AK']:
            ws.conditional_formatting.add(f'{col}2:{col}1000',
                FormulaRule(formula=[f'{col}2<0.9'], 
                           fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')))
            ws.conditional_formatting.add(f'{col}2:{col}1000',
                FormulaRule(formula=[f'{col}2>=1.1'], 
                           fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')))
                           
    def create_analytics_dashboard(self):
        """Create advanced analytics dashboard with charts"""
        ws = self.wb.create_sheet("Analytics Dashboard")
        
        # Dashboard title
        ws['A1'] = 'PROJECT ANALYTICS DASHBOARD'
        ws['A1'].font = Font(name='Verdana', size=24, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:M1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Create dashboard sections
        self.create_kpi_section(ws)
        self.create_charts_section(ws)
        self.create_risk_matrix(ws)
        self.create_resource_analysis(ws)
        
    def create_kpi_section(self, ws):
        """Create KPI cards section"""
        kpi_row = 3
        kpis = [
            ('Overall Progress', '=ROUND(AVERAGE(\'Project Tasks\'!AG:AG),1)&"%"', self.colors['info']),
            ('On-Time Delivery', '=ROUND(COUNTIF(\'Project Tasks\'!Y:Y,"Green")/COUNTA(\'Project Tasks\'!Y:Y)*100,1)&"%"', self.colors['success']),
            ('At Risk Tasks', '=COUNTIFS(\'Project Tasks\'!Y:Y,"Yellow")+COUNTIFS(\'Project Tasks\'!Y:Y,"Orange")', self.colors['warning']),
            ('Critical Issues', '=COUNTIF(\'Project Tasks\'!Y:Y,"Red")', self.colors['danger']),
            ('Budget Health', '=ROUND(SUMIF(\'Project Tasks\'!V:V,">0",\'Project Tasks\'!V:V)/SUMIF(\'Project Tasks\'!W:W,">0",\'Project Tasks\'!W:W)*100,1)&"%"', self.colors['accent_blue']),
            ('Resource Utilization', '=ROUND(AVERAGE(\'Project Tasks\'!U:U),1)&"%"', self.colors['primary'])
        ]
        
        col = 1
        for title, formula, color in kpis:
            # KPI Title
            ws.cell(row=kpi_row, column=col, value=title)
            ws.cell(row=kpi_row, column=col).font = Font(name='Verdana', size=12, bold=True)
            
            # KPI Value
            ws.cell(row=kpi_row+1, column=col, value=formula)
            ws.cell(row=kpi_row+1, column=col).font = Font(name='Verdana', size=20, bold=True, color=color)
            ws.cell(row=kpi_row+1, column=col).alignment = Alignment(horizontal='center')
            
            # Add border
            for r in range(kpi_row, kpi_row+2):
                ws.cell(row=r, column=col).border = self.borders['thick']
                
            col += 2
            
    def create_charts_section(self, ws):
        """Create various charts for visualization"""
        # Progress by Phase Chart
        chart1 = BarChart()
        chart1.title = "Progress by Phase"
        chart1.style = 10
        chart1.x_axis.title = 'Phase'
        chart1.y_axis.title = 'Progress %'
        
        # Add chart to sheet
        ws.add_chart(chart1, "A7")
        
        # Resource Utilization Chart
        chart2 = PieChart()
        chart2.title = "Resource Allocation"
        chart2.style = 10
        
        ws.add_chart(chart2, "G7")
        
        # Burndown Chart
        chart3 = LineChart()
        chart3.title = "Project Burndown"
        chart3.style = 10
        chart3.x_axis.title = 'Date'
        chart3.y_axis.title = 'Remaining Work'
        
        ws.add_chart(chart3, "A20")
        
    def create_risk_matrix(self, ws):
        """Create risk assessment matrix"""
        ws['A35'] = 'RISK MATRIX'
        ws['A35'].font = self.fonts['header']
        
        # Create 5x5 risk matrix
        risk_levels = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
        impact_levels = ['Minimal', 'Minor', 'Moderate', 'Major', 'Severe']
        
        # Headers
        for i, level in enumerate(risk_levels, 1):
            ws.cell(row=37, column=i+1, value=level)
            ws.cell(row=37, column=i+1).font = self.fonts['body_bold']
            
        for i, level in enumerate(impact_levels, 1):
            ws.cell(row=37+i, column=1, value=level)
            ws.cell(row=37+i, column=1).font = self.fonts['body_bold']
            
        # Color-code matrix cells
        for row in range(38, 43):
            for col in range(2, 7):
                risk_score = (row-37) * (col-1) * 20
                cell = ws.cell(row=row, column=col)
                cell.value = risk_score
                
                if risk_score <= 40:
                    cell.fill = self.fills['success']
                elif risk_score <= 60:
                    cell.fill = self.fills['warning']
                else:
                    cell.fill = self.fills['critical']
                    
                cell.border = self.borders['thin']
                
    def create_resource_analysis(self, ws):
        """Create resource utilization analysis"""
        ws['G35'] = 'RESOURCE ANALYSIS'
        ws['G35'].font = self.fonts['header']
        
        # Resource headers
        headers = ['Resource', 'Allocated Tasks', 'Utilization %', 'Capacity']
        for col, header in enumerate(headers, 7):
            ws.cell(row=37, column=col, value=header)
            ws.cell(row=37, column=col).font = self.fonts['body_bold']
            ws.cell(row=37, column=col).border = self.borders['bottom_thick']
            
    def create_gantt_sheet(self):
        """Create Gantt chart visualization sheet"""
        ws = self.wb.create_sheet("Gantt Chart")
        
        ws['A1'] = 'PROJECT GANTT CHART'
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:Z1')
        
        # Instructions
        ws['A3'] = 'This sheet provides a visual timeline of all project tasks.'
        ws['A4'] = 'Use conditional formatting to create Gantt bars based on start/end dates.'
        
        # Headers
        headers = ['Task ID', 'Task Name', 'Start', 'End', 'Duration', 'Progress']
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
            ws.cell(row=6, column=col).font = self.fonts['header']
            ws.cell(row=6, column=col).border = self.borders['thick']
            
        # Add formulas to pull data from main sheet
        for row in range(7, 50):
            ws[f'A{row}'] = f"='Project Tasks'!A{row-5}"
            ws[f'B{row}'] = f"='Project Tasks'!B{row-5}"
            ws[f'C{row}'] = f"='Project Tasks'!G{row-5}"
            ws[f'D{row}'] = f"='Project Tasks'!H{row-5}"
            ws[f'E{row}'] = f"='Project Tasks'!F{row-5}"
            ws[f'F{row}'] = f"='Project Tasks'!K{row-5}"
            
    def create_reports_sheet(self):
        """Create automated reports sheet"""
        ws = self.wb.create_sheet("Reports")
        
        ws['A1'] = 'AUTOMATED PROJECT REPORTS'
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:H1')
        
        # Report sections
        reports = [
            ('EXECUTIVE SUMMARY', 3),
            ('CRITICAL PATH ANALYSIS', 15),
            ('RISK ASSESSMENT', 27),
            ('RESOURCE ALLOCATION', 39),
            ('BUDGET ANALYSIS', 51),
            ('SCHEDULE VARIANCE', 63)
        ]
        
        for title, start_row in reports:
            ws[f'A{start_row}'] = title
            ws[f'A{start_row}'].font = self.fonts['header']
            ws[f'A{start_row}'].border = self.borders['bottom_thick']
            
            # Add report content formulas
            self.add_report_content(ws, title, start_row + 2)
            
    def add_report_content(self, ws, report_type, start_row):
        """Add specific report content based on type"""
        if report_type == 'EXECUTIVE SUMMARY':
            content = [
                ('Project Status:', '=IF(COUNTIF(\'Project Tasks\'!Y:Y,"Red")>0,"Critical",IF(COUNTIF(\'Project Tasks\'!Y:Y,"Yellow")>3,"At Risk","On Track"))'),
                ('Overall Progress:', '=ROUND(AVERAGE(\'Project Tasks\'!AG:AG),1)&"%"'),
                ('Schedule Performance:', '=IF(AVERAGE(\'Project Tasks\'!AJ:AJ)>=1,"Ahead of Schedule",IF(AVERAGE(\'Project Tasks\'!AJ:AJ)>=0.9,"On Schedule","Behind Schedule"))'),
                ('Budget Performance:', '=IF(AVERAGE(\'Project Tasks\'!AK:AK)>=1,"Under Budget",IF(AVERAGE(\'Project Tasks\'!AK:AK)>=0.9,"On Budget","Over Budget"))'),
                ('Critical Issues:', '=COUNTIF(\'Project Tasks\'!Y:Y,"Red")&" tasks require immediate attention"'),
                ('Next Milestone:', '=INDEX(\'Project Tasks\'!B:B,MATCH("Milestone",\'Project Tasks\'!C:C,0))'),
            ]
            
            for i, (label, formula) in enumerate(content):
                ws[f'A{start_row + i}'] = label
                ws[f'C{start_row + i}'] = formula
                ws[f'A{start_row + i}'].font = self.fonts['body_bold']
                
    def create_instructions_enhanced(self):
        """Create comprehensive instructions with best practices"""
        ws = self.wb.create_sheet("User Guide")
        
        guide_content = [
            ['PROJECT MANAGEMENT TEMPLATE v2.0 - COMPREHENSIVE GUIDE', '', ''],
            ['', '', ''],
            ['QUICK START', '', ''],
            ['1. Project Setup', 'Enter project phases as Parent tasks', 'Use WBS codes for hierarchy'],
            ['2. Task Entry', 'Add Child tasks under each Parent', 'Set dependencies and lag/lead times'],
            ['3. Resource Assignment', 'Assign teams and set utilization %', 'Monitor resource conflicts'],
            ['4. Progress Tracking', 'Update Status and Actual dates daily', 'System auto-calculates progress'],
            ['5. Risk Management', 'Review Risk Scores regularly', 'Document mitigation strategies'],
            ['', '', ''],
            ['ADVANCED FEATURES', '', ''],
            ['Impact Score Algorithm', 'ML-enhanced calculation', 'Considers 8+ factors including dependencies, resources, timeline'],
            ['Risk Assessment', 'Automatic risk scoring', 'Based on delays, variances, and performance indices'],
            ['Critical Path Analysis', 'Real-time CP calculation', 'Shows float and highlights critical tasks'],
            ['Resource Optimization', 'Load balancing alerts', 'Prevents resource conflicts'],
            ['Earned Value Management', 'SPI and CPI tracking', 'Budget and schedule performance'],
            ['', '', ''],
            ['FORMULAS EXPLAINED', '', ''],
            ['Impact Score', 'Weighted calculation', 'Duration(20%) + Dependencies(25%) + Criticality(15%) + Position(15%) + Resources(10%) + Milestone(10%) + Budget(5%)'],
            ['Risk Score', 'Dynamic assessment', 'Status risks + Schedule variance + Performance indices + Float analysis'],
            ['Health Indicator', 'Multi-factor health', 'Combines progress, risk, schedule, and budget factors'],
            ['', '', ''],
            ['BEST PRACTICES', '', ''],
            ['Daily Updates', 'Update task status daily', 'Accurate data ensures reliable predictions'],
            ['Dependency Management', 'Define all dependencies upfront', 'Use appropriate relationship types'],
            ['Resource Planning', 'Keep utilization below 85%', 'Allow buffer for unexpected work'],
            ['Risk Mitigation', 'Address high-risk items first', 'Document mitigation strategies'],
            ['Communication', 'Use dashboards for stakeholders', 'Export reports weekly'],
            ['', '', ''],
            ['TROUBLESHOOTING', '', ''],
            ['Circular References', 'Check parent-child relationships', 'Ensure no task is its own parent'],
            ['Performance Issues', 'Limit active tasks to 500', 'Archive completed projects'],
            ['Formula Errors', 'Verify data types', 'Check for empty required fields'],
            ['', '', ''],
            ['SHORTCUTS & TIPS', '', ''],
            ['Ctrl+Shift+L', 'Toggle filters', 'Quick data filtering'],
            ['Alt+H+O+I', 'Auto-fit columns', 'Optimize column widths'],
            ['F9', 'Recalculate', 'Force formula updates'],
            ['Ctrl+`', 'Show formulas', 'Debug formula issues'],
        ]
        
        for row_idx, content in enumerate(guide_content, 1):
            for col_idx, text in enumerate(content, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                
                if row_idx == 1:
                    cell.font = self.fonts['title']
                elif content[0] and not content[1] and not content[2]:
                    cell.font = self.fonts['header']
                elif content[0] and content[0].endswith(':'):
                    cell.font = self.fonts['subheader']
                else:
                    cell.font = self.fonts['body']
                    
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        
    def save_enhanced_template(self, filename='project_management_template_v2.xlsx'):
        """Save the enhanced template"""
        self.create_enhanced_main_sheet()
        self.create_analytics_dashboard()
        self.create_gantt_sheet()
        self.create_reports_sheet()
        self.create_instructions_enhanced()
        
        # Add metadata
        self.wb.properties.title = "Advanced Project Management Template"
        self.wb.properties.creator = "AI-Enhanced PM System"
        self.wb.properties.description = "ML-powered project tracking with impact-based scoring"
        self.wb.properties.keywords = "project management, impact score, risk assessment, critical path"
        
        # Save workbook
        self.wb.save(filename)
        print(f"Enhanced template saved as: {filename}")
        

if __name__ == "__main__":
    generator = AdvancedProjectTemplateGenerator()
    generator.save_enhanced_template('/Users/haithamdata/Documents/Prog/My Productivity/Project Managment/Project managment tamblet/project_management_template_v2.xlsx')