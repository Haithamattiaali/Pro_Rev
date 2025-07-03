#!/usr/bin/env python3
"""
Project Management Template Generator - Working Version
Creates a fully functional Excel template with calculated values
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

class WorkingProjectTemplate:
    def __init__(self):
        self.wb = Workbook()
        self.data = []
        self.setup_styles()
        
    def setup_styles(self):
        """Define color scheme and styles"""
        self.colors = {
            'primary': '9E1F63',
            'secondary': '721548',
            'accent_blue': '005B8C',
            'success': '27AE60',
            'warning': 'F39C12',
            'danger': 'E74C3C',
            'light': 'ECF0F1',
            'white': 'FFFFFF'
        }
        
        self.fonts = {
            'header': Font(name='Verdana', size=12, bold=True, color=self.colors['primary']),
            'body': Font(name='Verdana', size=10),
            'body_bold': Font(name='Verdana', size=10, bold=True)
        }
        
        self.fills = {
            'header': PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid'),
            'parent': PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid'),
            'success': PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid'),
            'warning': PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid'),
            'danger': PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid')
        }
        
        self.borders = {
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
    def calculate_impact_score(self, row_data):
        """Calculate impact score based on multiple factors"""
        try:
            # Extract values
            duration = float(row_data.get('Duration (Days)', 0))
            task_type = row_data.get('Task Type', '')
            criticality = row_data.get('Criticality Level', 'Medium')
            dependencies = str(row_data.get('Dependencies', ''))
            blocking_tasks = str(row_data.get('Blocking Tasks', ''))
            agility = row_data.get('Task Agility', '')
            status = row_data.get('Status', '')
            milestone = row_data.get('Milestone', 'No')
            budget = float(row_data.get('Cost Budget', 0))
            
            # Base calculations
            duration_weight = min(duration / 40 * 25, 25)  # Normalize to max 25
            
            # Dependency count
            dep_count = 0
            if dependencies and dependencies not in ['', '0']:
                dep_count = len(dependencies.split(','))
            dependency_weight = min(dep_count * 5, 30)
            
            # Criticality weight
            crit_weights = {'Critical': 20, 'High': 15, 'Medium': 10, 'Low': 5, 'Minimal': 2}
            criticality_weight = crit_weights.get(criticality, 10)
            
            # Timeline position weight (earlier = higher)
            timeline_weight = 15  # Default for mid-project
            
            # Milestone bonus
            milestone_bonus = 15 if milestone == 'Yes' else 0
            
            # Budget impact (logarithmic)
            budget_weight = 0
            if budget > 0:
                budget_weight = min(np.log10(budget + 1) / 6 * 10, 10)
            
            # Calculate base score
            base_score = (duration_weight + dependency_weight + criticality_weight + 
                         timeline_weight + milestone_bonus + budget_weight)
            
            # Apply multipliers
            # Blocking factor
            blocking_multiplier = 1.5 if blocking_tasks and blocking_tasks not in ['', '0'] else 1.0
            
            # Parallel reduction
            parallel_multiplier = 0.7 if agility == 'Parallel' else 1.0
            
            # Status multiplier
            status_multiplier = 1.0
            if status == 'Delayed':
                status_multiplier = 1.3
            elif status == 'Blocked':
                status_multiplier = 1.8
                
            # Final score
            impact_score = int(base_score * blocking_multiplier * parallel_multiplier * status_multiplier)
            return min(impact_score, 100)  # Cap at 100
            
        except Exception as e:
            return 50  # Default score on error
            
    def calculate_risk_score(self, row_data):
        """Calculate risk score based on multiple factors"""
        try:
            status = row_data.get('Status', '')
            progress = float(row_data.get('% Complete', 0))
            criticality = row_data.get('Criticality Level', 'Medium')
            health = row_data.get('Health Indicator', 'Green')
            
            # Base risk from status
            status_risk = 0
            if status == 'Blocked':
                status_risk = 30
            elif status == 'Delayed':
                status_risk = 20
            elif status == 'On Hold':
                status_risk = 15
                
            # Progress risk
            progress_risk = 0
            if progress < 50 and status == 'In Progress':
                progress_risk = 20
            elif progress < 80 and status == 'In Progress':
                progress_risk = 10
                
            # Health risk
            health_risk = 0
            if health == 'Red':
                health_risk = 25
            elif health == 'Orange':
                health_risk = 15
            elif health == 'Yellow':
                health_risk = 5
                
            # Criticality multiplier
            crit_mult = {'Critical': 1.5, 'High': 1.2, 'Medium': 1.0, 'Low': 0.8, 'Minimal': 0.5}
            criticality_multiplier = crit_mult.get(criticality, 1.0)
            
            # Calculate total risk
            total_risk = (status_risk + progress_risk + health_risk) * criticality_multiplier
            return min(int(total_risk), 100)
            
        except Exception:
            return 0
            
    def calculate_health_indicator(self, row_data):
        """Determine health indicator based on multiple factors"""
        try:
            status = row_data.get('Status', '')
            progress = float(row_data.get('% Complete', 0))
            risk_score = row_data.get('Risk Score', 0)
            
            if status == 'Complete':
                return 'Green'
            elif status == 'Cancelled':
                return 'Black'
            elif status == 'Blocked' or risk_score > 80:
                return 'Red'
            elif status == 'Delayed' or risk_score > 60:
                return 'Orange'
            elif risk_score > 40 or (progress < 80 and status == 'In Progress'):
                return 'Yellow'
            else:
                return 'Green'
                
        except Exception:
            return 'Yellow'
            
    def calculate_priority_score(self, impact, risk):
        """Calculate priority score combining impact and risk"""
        try:
            return int(impact * 0.6 + risk * 0.4)
        except:
            return 50
            
    def create_sample_data(self):
        """Create comprehensive sample project data"""
        # Define sample tasks
        tasks = [
            # Project root
            {
                'Task ID': 'P001',
                'Task Name': 'Digital Transformation Initiative',
                'Task Type': 'Parent',
                'Parent Task ID': '',
                'WBS Code': '1',
                'Duration (Days)': 120,
                'Start Date': '2024-01-01',
                'End Date': '',
                'Actual Start': '2024-01-01',
                'Actual End': '',
                '% Complete': 35,
                'Task Agility': 'Sequential',
                'Dependencies': '',
                'Dependency Type': '',
                'Lag/Lead': 0,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Program Management',
                'Resource Load %': 100,
                'Cost Budget': 5000000,
                'Actual Cost': 1750000,
                'Status': 'In Progress',
                'Milestone': 'No',
                'Deliverables': 'Complete digital transformation',
                'Blocking Tasks': '',
                'Blocked By': '',
                'Notes': 'Strategic initiative for 2024'
            },
            
            # Phase 1
            {
                'Task ID': 'P002',
                'Task Name': 'Phase 1: Discovery & Planning',
                'Task Type': 'Parent',
                'Parent Task ID': 'P001',
                'WBS Code': '1.1',
                'Duration (Days)': 20,
                'Start Date': '2024-01-01',
                'End Date': '',
                'Actual Start': '2024-01-01',
                'Actual End': '2024-01-20',
                '% Complete': 100,
                'Task Agility': 'Sequential',
                'Dependencies': '',
                'Dependency Type': '',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'PM Team',
                'Resource Load %': 100,
                'Cost Budget': 500000,
                'Actual Cost': 485000,
                'Status': 'Complete',
                'Milestone': 'No',
                'Deliverables': 'Requirements and architecture',
                'Blocking Tasks': 'P003,P004',
                'Blocked By': '',
                'Notes': 'Foundation phase completed successfully'
            },
            
            # Child tasks for Phase 1
            {
                'Task ID': 'C001',
                'Task Name': 'Stakeholder Interviews',
                'Task Type': 'Child',
                'Parent Task ID': 'P002',
                'WBS Code': '1.1.1',
                'Duration (Days)': 5,
                'Start Date': '2024-01-02',
                'End Date': '',
                'Actual Start': '2024-01-02',
                'Actual End': '2024-01-06',
                '% Complete': 100,
                'Task Agility': 'Parallel',
                'Dependencies': '',
                'Dependency Type': '',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'BA Team',
                'Resource Load %': 80,
                'Cost Budget': 50000,
                'Actual Cost': 48000,
                'Status': 'Complete',
                'Milestone': 'No',
                'Deliverables': 'Interview notes and findings',
                'Blocking Tasks': 'C003',
                'Blocked By': '',
                'Notes': 'Excellent insights gathered'
            },
            
            {
                'Task ID': 'C002',
                'Task Name': 'Current State Analysis',
                'Task Type': 'Child',
                'Parent Task ID': 'P002',
                'WBS Code': '1.1.2',
                'Duration (Days)': 5,
                'Start Date': '2024-01-02',
                'End Date': '',
                'Actual Start': '2024-01-02',
                'Actual End': '2024-01-06',
                '% Complete': 100,
                'Task Agility': 'Parallel',
                'Dependencies': '',
                'Dependency Type': '',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'Tech Architects',
                'Resource Load %': 100,
                'Cost Budget': 75000,
                'Actual Cost': 72000,
                'Status': 'Complete',
                'Milestone': 'No',
                'Deliverables': 'As-is documentation',
                'Blocking Tasks': 'C003',
                'Blocked By': '',
                'Notes': 'Technical debt identified'
            },
            
            {
                'Task ID': 'C003',
                'Task Name': 'Future State Design',
                'Task Type': 'Child',
                'Parent Task ID': 'P002',
                'WBS Code': '1.1.3',
                'Duration (Days)': 7,
                'Start Date': '2024-01-07',
                'End Date': '',
                'Actual Start': '2024-01-07',
                'Actual End': '2024-01-13',
                '% Complete': 100,
                'Task Agility': 'Sequential',
                'Dependencies': 'C001,C002',
                'Dependency Type': 'FS',
                'Lag/Lead': 0,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Solution Architects',
                'Resource Load %': 100,
                'Cost Budget': 100000,
                'Actual Cost': 98000,
                'Status': 'Complete',
                'Milestone': 'No',
                'Deliverables': 'To-be architecture',
                'Blocking Tasks': 'C004',
                'Blocked By': 'C001,C002',
                'Notes': 'Microservices architecture approved'
            },
            
            # Phase 2
            {
                'Task ID': 'P003',
                'Task Name': 'Phase 2: Infrastructure Setup',
                'Task Type': 'Parent',
                'Parent Task ID': 'P001',
                'WBS Code': '1.2',
                'Duration (Days)': 30,
                'Start Date': '2024-01-21',
                'End Date': '',
                'Actual Start': '2024-01-21',
                'Actual End': '',
                '% Complete': 65,
                'Task Agility': 'Parallel',
                'Dependencies': 'P002',
                'Dependency Type': 'FS',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'Infrastructure Team',
                'Resource Load %': 100,
                'Cost Budget': 1500000,
                'Actual Cost': 975000,
                'Status': 'In Progress',
                'Milestone': 'No',
                'Deliverables': 'Cloud infrastructure ready',
                'Blocking Tasks': 'P004',
                'Blocked By': 'P002',
                'Notes': 'AWS multi-region setup'
            },
            
            # Infrastructure child tasks
            {
                'Task ID': 'C005',
                'Task Name': 'Cloud Account Setup',
                'Task Type': 'Child',
                'Parent Task ID': 'P003',
                'WBS Code': '1.2.1',
                'Duration (Days)': 3,
                'Start Date': '2024-01-21',
                'End Date': '',
                'Actual Start': '2024-01-21',
                'Actual End': '2024-01-23',
                '% Complete': 100,
                'Task Agility': 'Sequential',
                'Dependencies': 'P002',
                'Dependency Type': 'FS',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'Cloud Team',
                'Resource Load %': 50,
                'Cost Budget': 10000,
                'Actual Cost': 9500,
                'Status': 'Complete',
                'Milestone': 'No',
                'Deliverables': 'AWS accounts ready',
                'Blocking Tasks': 'C006,C007',
                'Blocked By': '',
                'Notes': 'Multi-account strategy implemented'
            },
            
            {
                'Task ID': 'C007',
                'Task Name': 'Security Framework',
                'Task Type': 'Child',
                'Parent Task ID': 'P003',
                'WBS Code': '1.2.3',
                'Duration (Days)': 10,
                'Start Date': '2024-01-24',
                'End Date': '',
                'Actual Start': '2024-01-24',
                'Actual End': '',
                '% Complete': 70,
                'Task Agility': 'Parallel',
                'Dependencies': 'C005',
                'Dependency Type': 'FS',
                'Lag/Lead': 0,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Security Team',
                'Resource Load %': 100,
                'Cost Budget': 200000,
                'Actual Cost': 140000,
                'Status': 'In Progress',
                'Milestone': 'No',
                'Deliverables': 'Security policies & tools',
                'Blocking Tasks': '',
                'Blocked By': 'C005',
                'Notes': 'GDPR compliance in progress'
            },
            
            # Phase 3
            {
                'Task ID': 'P004',
                'Task Name': 'Phase 3: Development Sprint 1',
                'Task Type': 'Parent',
                'Parent Task ID': 'P001',
                'WBS Code': '1.3',
                'Duration (Days)': 40,
                'Start Date': '2024-02-20',
                'End Date': '',
                'Actual Start': '',
                'Actual End': '',
                '% Complete': 0,
                'Task Agility': 'Parallel',
                'Dependencies': 'P003',
                'Dependency Type': 'SS',
                'Lag/Lead': 10,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Dev Team Alpha',
                'Resource Load %': 100,
                'Cost Budget': 1000000,
                'Actual Cost': 0,
                'Status': 'Planning',
                'Milestone': 'No',
                'Deliverables': 'Core modules developed',
                'Blocking Tasks': 'P005',
                'Blocked By': 'P003',
                'Notes': 'Agile development approach'
            },
            
            # Milestones
            {
                'Task ID': 'M001',
                'Task Name': 'Project Kickoff Complete',
                'Task Type': 'Milestone',
                'Parent Task ID': 'P002',
                'WBS Code': '1.1.0',
                'Duration (Days)': 0,
                'Start Date': '2024-01-01',
                'End Date': '',
                'Actual Start': '2024-01-01',
                'Actual End': '2024-01-01',
                '% Complete': 100,
                'Task Agility': 'Sequential',
                'Dependencies': '',
                'Dependency Type': '',
                'Lag/Lead': 0,
                'Criticality Level': 'High',
                'Resource Assignment': 'All Teams',
                'Resource Load %': 0,
                'Cost Budget': 0,
                'Actual Cost': 0,
                'Status': 'Complete',
                'Milestone': 'Yes',
                'Deliverables': 'Kickoff meeting held',
                'Blocking Tasks': '',
                'Blocked By': '',
                'Notes': 'Successful project launch'
            },
            
            {
                'Task ID': 'M002',
                'Task Name': 'Infrastructure Ready',
                'Task Type': 'Milestone',
                'Parent Task ID': 'P003',
                'WBS Code': '1.2.0',
                'Duration (Days)': 0,
                'Start Date': '2024-02-19',
                'End Date': '',
                'Actual Start': '',
                'Actual End': '',
                '% Complete': 0,
                'Task Agility': 'Sequential',
                'Dependencies': 'C007',
                'Dependency Type': 'FS',
                'Lag/Lead': 0,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Infrastructure Team',
                'Resource Load %': 0,
                'Cost Budget': 0,
                'Actual Cost': 0,
                'Status': 'Not Started',
                'Milestone': 'Yes',
                'Deliverables': 'Infrastructure sign-off',
                'Blocking Tasks': 'P004',
                'Blocked By': '',
                'Notes': 'Gate 2 milestone'
            },
            
            # Risk item
            {
                'Task ID': 'R001',
                'Task Name': 'Security Vulnerability Assessment',
                'Task Type': 'Child',
                'Parent Task ID': 'P003',
                'WBS Code': '1.2.9',
                'Duration (Days)': 3,
                'Start Date': '2024-02-01',
                'End Date': '',
                'Actual Start': '2024-02-01',
                'Actual End': '',
                '% Complete': 40,
                'Task Agility': 'Parallel',
                'Dependencies': 'C007',
                'Dependency Type': 'SS',
                'Lag/Lead': 3,
                'Criticality Level': 'Critical',
                'Resource Assignment': 'Security Team',
                'Resource Load %': 100,
                'Cost Budget': 30000,
                'Actual Cost': 12000,
                'Status': 'Delayed',
                'Milestone': 'No',
                'Deliverables': 'Vulnerability report',
                'Blocking Tasks': '',
                'Blocked By': '',
                'Notes': 'Critical vulnerabilities found - remediation in progress'
            }
        ]
        
        # Calculate additional fields
        for task in tasks:
            # Calculate end dates
            if task['Start Date'] and task['Duration (Days)'] > 0:
                start = pd.to_datetime(task['Start Date'])
                # Add business days
                end = pd.bdate_range(start, periods=task['Duration (Days)'] + 1)[-1]
                task['End Date'] = end.strftime('%Y-%m-%d')
            
            # Calculate impact score
            task['Impact Score'] = self.calculate_impact_score(task)
            
            # Calculate risk score
            task['Risk Score'] = self.calculate_risk_score(task)
            
            # Calculate health indicator
            task['Health Indicator'] = self.calculate_health_indicator(task)
            
            # Calculate priority score
            task['Priority Score'] = self.calculate_priority_score(
                task['Impact Score'], 
                task['Risk Score']
            )
            
            # Calculate float (simplified)
            task['Total Float'] = 0 if task['Criticality Level'] == 'Critical' else 5
            task['Free Float'] = task['Total Float']
            
            # Critical path
            task['Critical Path'] = 'Yes' if task['Total Float'] == 0 else 'No'
            
            # Progress calculations
            if task['Task Type'] == 'Parent':
                task['Weighted Progress'] = task['% Complete']  # Simplified
            else:
                task['Weighted Progress'] = task['% Complete']
            
            task['Rolled Up Progress'] = task['% Complete']
            
            # Variance (simplified)
            task['Variance Days'] = 0 if task['Status'] == 'Complete' else -2 if task['Status'] == 'Delayed' else 0
            
            # Performance indices
            task['SPI'] = 1.0 if task['% Complete'] >= 50 else 0.8
            task['CPI'] = 1.0 if task['Actual Cost'] <= task['Cost Budget'] * 0.5 else 0.9
            
            # Risk mitigation
            if task['Risk Score'] > 60:
                task['Risk Mitigation'] = 'Immediate action required'
            elif task['Risk Score'] > 40:
                task['Risk Mitigation'] = 'Monitor closely'
            else:
                task['Risk Mitigation'] = ''
                
            # Lessons learned
            task['Lessons Learned'] = 'In progress' if task['Status'] == 'Complete' else ''
            
        self.data = tasks
        return tasks
        
    def create_main_sheet(self):
        """Create the main project tracking sheet"""
        ws = self.wb.active
        ws.title = "Project Tasks"
        
        # Get sample data
        tasks = self.create_sample_data()
        
        # Define headers
        headers = list(tasks[0].keys())
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.borders['thin']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_idx, task in enumerate(tasks, 2):
            for col_idx, header in enumerate(headers, 1):
                value = task.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.borders['thin']
                
                # Apply formatting
                if task['Task Type'] == 'Parent':
                    cell.fill = self.fills['parent']
                    cell.font = self.fonts['body_bold']
                else:
                    cell.font = self.fonts['body']
                    
                # Format numbers
                if header in ['% Complete', 'Resource Load %']:
                    cell.number_format = '0%'
                    cell.value = value / 100 if value else 0
                elif header in ['Cost Budget', 'Actual Cost']:
                    cell.number_format = '$#,##0'
                elif header in ['SPI', 'CPI']:
                    cell.number_format = '0.00'
                    
                # Color health indicators
                if header == 'Health Indicator':
                    if value == 'Green':
                        cell.fill = self.fills['success']
                    elif value == 'Yellow':
                        cell.fill = self.fills['warning']
                    elif value in ['Red', 'Orange']:
                        cell.fill = self.fills['danger']
                        
        # Set column widths
        self.set_column_widths(ws, headers)
        
        # Add conditional formatting
        self.add_conditional_formatting(ws, len(tasks) + 1)
        
        # Freeze panes
        ws.freeze_panes = 'E2'
        
    def set_column_widths(self, ws, headers):
        """Set appropriate column widths"""
        width_map = {
            'Task ID': 10,
            'Task Name': 35,
            'Task Type': 12,
            'Parent Task ID': 12,
            'WBS Code': 10,
            'Duration (Days)': 12,
            'Start Date': 12,
            'End Date': 12,
            'Actual Start': 12,
            'Actual End': 12,
            '% Complete': 12,
            'Task Agility': 12,
            'Dependencies': 15,
            'Dependency Type': 12,
            'Lag/Lead': 10,
            'Impact Score': 12,
            'Risk Score': 12,
            'Criticality Level': 14,
            'Priority Score': 12,
            'Resource Assignment': 20,
            'Resource Load %': 14,
            'Cost Budget': 12,
            'Actual Cost': 12,
            'Status': 12,
            'Health Indicator': 14,
            'Milestone': 10,
            'Deliverables': 25,
            'Blocking Tasks': 15,
            'Blocked By': 15,
            'Total Float': 10,
            'Free Float': 10,
            'Critical Path': 12,
            'Weighted Progress': 15,
            'Rolled Up Progress': 15,
            'Variance Days': 12,
            'SPI': 8,
            'CPI': 8,
            'Risk Mitigation': 20,
            'Lessons Learned': 20,
            'Notes': 30
        }
        
        for col, header in enumerate(headers, 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width_map.get(header, 12)
            
    def add_conditional_formatting(self, ws, max_row):
        """Add conditional formatting for visual enhancement"""
        # Progress bars
        progress_col = None
        impact_col = None
        risk_col = None
        
        # Find column indices
        for cell in ws[1]:
            if cell.value == '% Complete':
                progress_col = cell.column
            elif cell.value == 'Impact Score':
                impact_col = cell.column
            elif cell.value == 'Risk Score':
                risk_col = cell.column
                
        # Add color scales
        if progress_col:
            ws.conditional_formatting.add(
                f'{get_column_letter(progress_col)}2:{get_column_letter(progress_col)}{max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF0000',
                    mid_type='num', mid_value=0.5, mid_color='FFFF00',
                    end_type='num', end_value=1, end_color='00FF00'
                )
            )
            
        if impact_col:
            ws.conditional_formatting.add(
                f'{get_column_letter(impact_col)}2:{get_column_letter(impact_col)}{max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FFFFFF',
                    mid_type='num', mid_value=50, mid_color='FFA500',
                    end_type='num', end_value=100, end_color='FF0000'
                )
            )
            
        if risk_col:
            ws.conditional_formatting.add(
                f'{get_column_letter(risk_col)}2:{get_column_letter(risk_col)}{max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='00FF00',
                    mid_type='num', mid_value=50, mid_color='FFFF00',
                    end_type='num', end_value=100, end_color='FF0000'
                )
            )
            
    def create_dashboard_sheet(self):
        """Create a simple dashboard sheet"""
        ws = self.wb.create_sheet("Dashboard")
        
        # Title
        ws['A1'] = 'PROJECT DASHBOARD'
        ws['A1'].font = Font(name='Verdana', size=20, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:F1')
        
        # Calculate KPIs from data
        total_tasks = len(self.data)
        complete_tasks = len([t for t in self.data if t['Status'] == 'Complete'])
        at_risk_tasks = len([t for t in self.data if t['Risk Score'] > 60])
        total_budget = sum(t['Cost Budget'] for t in self.data)
        actual_cost = sum(t['Actual Cost'] for t in self.data)
        
        # KPI Section
        kpis = [
            ('Total Tasks:', total_tasks),
            ('Completed:', complete_tasks),
            ('At Risk:', at_risk_tasks),
            ('Total Budget:', f'${total_budget:,.0f}'),
            ('Actual Cost:', f'${actual_cost:,.0f}'),
            ('Budget Used:', f'{(actual_cost/total_budget*100):.1f}%' if total_budget > 0 else '0%')
        ]
        
        row = 3
        for label, value in kpis:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = self.fonts['body_bold']
            ws[f'C{row}'].font = Font(name='Verdana', size=12, bold=True)
            row += 1
            
        # Status Summary
        ws['A10'] = 'STATUS SUMMARY'
        ws['A10'].font = self.fonts['header']
        
        status_counts = {}
        for task in self.data:
            status = task['Status']
            status_counts[status] = status_counts.get(status, 0) + 1
            
        row = 12
        for status, count in status_counts.items():
            ws[f'A{row}'] = status
            ws[f'C{row}'] = count
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['C'].width = 15
        
    def create_instructions_sheet(self):
        """Create instructions sheet"""
        ws = self.wb.create_sheet("Instructions")
        
        instructions = [
            ['WORKING PROJECT MANAGEMENT TEMPLATE', ''],
            ['', ''],
            ['OVERVIEW', 'This template provides a comprehensive project tracking system with calculated impact and risk scores.'],
            ['', ''],
            ['KEY FEATURES:', ''],
            ['Impact Score', 'Automatically calculated based on duration, dependencies, criticality, and other factors'],
            ['Risk Score', 'Dynamic assessment based on status, progress, and health indicators'],
            ['Health Indicator', 'Visual status using color coding (Green/Yellow/Orange/Red)'],
            ['Priority Score', 'Combined impact and risk for prioritization'],
            ['', ''],
            ['USING THE TEMPLATE:', ''],
            ['1. Data Entry', 'Enter your project tasks with parent-child relationships'],
            ['2. Update Progress', 'Regularly update % Complete and Status fields'],
            ['3. Monitor Health', 'Watch the Health Indicator and Risk Score columns'],
            ['4. Review Dashboard', 'Check the Dashboard sheet for KPIs'],
            ['', ''],
            ['FORMULAS:', ''],
            ['Impact Score', 'Considers duration, dependencies, criticality, blocking tasks, and budget'],
            ['Risk Score', 'Based on status, progress delays, and health indicators'],
            ['Priority', 'Impact × 0.6 + Risk × 0.4'],
            ['', ''],
            ['BEST PRACTICES:', ''],
            ['Daily Updates', 'Update task status and progress daily'],
            ['Risk Review', 'Address high-risk items immediately'],
            ['Dependencies', 'Keep dependency information current'],
            ['Resource Load', 'Monitor resource utilization levels']
        ]
        
        for row_idx, (label, desc) in enumerate(instructions, 1):
            ws[f'A{row_idx}'] = label
            ws[f'B{row_idx}'] = desc
            
            if row_idx == 1:
                ws[f'A{row_idx}'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
            elif label in ['OVERVIEW', 'KEY FEATURES:', 'USING THE TEMPLATE:', 'FORMULAS:', 'BEST PRACTICES:']:
                ws[f'A{row_idx}'].font = self.fonts['header']
            else:
                ws[f'A{row_idx}'].font = self.fonts['body_bold']
                ws[f'B{row_idx}'].font = self.fonts['body']
                
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
        
    def create_syncup_dashboard_sheets(self):
        """Create comprehensive sync-up dashboard sheets"""
        # 1. Project Pulse Sheet
        self.create_project_pulse_sheet()
        
        # 2. Impact Matrix Sheet
        self.create_impact_matrix_sheet()
        
        # 3. Resource Orchestra Sheet
        self.create_resource_orchestra_sheet()
        
        # 4. Timeline Rhythm Sheet
        self.create_timeline_rhythm_sheet()
        
        # 5. Decision Command Sheet
        self.create_decision_command_sheet()
        
        # 6. Predictive Insights Sheet
        self.create_predictive_insights_sheet()
        
    def create_project_pulse_sheet(self):
        """Create project health pulse indicators"""
        ws = self.wb.create_sheet("Project Pulse")
        
        # Title
        ws['A1'] = 'PROJECT PULSE - HEALTH INDICATORS'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:F1')
        
        # Calculate overall health score
        total_tasks = len(self.data)
        complete_tasks = len([t for t in self.data if t['Status'] == 'Complete'])
        at_risk_tasks = len([t for t in self.data if t['Risk Score'] > 40])
        critical_tasks = len([t for t in self.data if t['Criticality Level'] == 'Critical'])
        
        avg_progress = np.mean([t['% Complete'] for t in self.data])
        avg_spi = np.mean([t['SPI'] for t in self.data if t['SPI'] > 0])
        avg_cpi = np.mean([t['CPI'] for t in self.data if t['CPI'] > 0])
        
        # Overall health calculation
        progress_score = avg_progress
        schedule_score = min(avg_spi * 100, 100) if avg_spi > 0 else 50
        budget_score = min(avg_cpi * 100, 100) if avg_cpi > 0 else 50
        risk_score = max(0, 100 - (at_risk_tasks / total_tasks * 100)) if total_tasks > 0 else 100
        
        overall_health = int((progress_score * 0.3 + schedule_score * 0.25 + 
                             budget_score * 0.25 + risk_score * 0.2))
        
        # Health metrics
        health_data = [
            ['Metric', 'Score', 'Status', 'Trend', 'Target', 'Gap'],
            ['Overall Project Health', overall_health, self.get_health_status(overall_health), '↑ +5%', 85, 85-overall_health],
            ['Schedule Performance', int(schedule_score), self.get_health_status(schedule_score), '↓ -2%', 90, 90-int(schedule_score)],
            ['Budget Performance', int(budget_score), self.get_health_status(budget_score), '→ 0%', 90, 90-int(budget_score)],
            ['Risk Management', int(risk_score), self.get_health_status(risk_score), '↑ +3%', 80, 80-int(risk_score)],
            ['Team Utilization', 78, 'Yellow', '↑ +5%', 75, -3],
        ]
        
        # Write health data
        for row_idx, row_data in enumerate(health_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header
                    cell.font = self.fonts['header']
                    cell.fill = self.fills['header']
                else:
                    cell.font = self.fonts['body']
                    # Color code status column
                    if col_idx == 3 and row_idx > 3:
                        if value == 'Green':
                            cell.fill = self.fills['success']
                        elif value == 'Yellow':
                            cell.fill = self.fills['warning']
                        elif value == 'Red':
                            cell.fill = self.fills['danger']
                cell.border = self.borders['thin']
                
        # Critical Alerts Section
        ws['A10'] = 'CRITICAL ALERTS'
        ws['A10'].font = self.fonts['header']
        ws.merge_cells('A10:F10')
        
        alerts = []
        # Check for blockers
        blocked_tasks = [t for t in self.data if t['Status'] == 'Blocked']
        if blocked_tasks:
            alerts.append(['BLOCKER', f"{len(blocked_tasks)} tasks blocked", 'Red', 'Immediate action required'])
            
        # Check for delays
        delayed_tasks = [t for t in self.data if t['Status'] == 'Delayed']
        if delayed_tasks:
            alerts.append(['DELAY', f"{len(delayed_tasks)} tasks delayed", 'Orange', 'Schedule at risk'])
            
        # Check for resource overload
        overloaded = [t for t in self.data if t.get('Resource Load %', 0) > 85]
        if overloaded:
            alerts.append(['RESOURCE', f"{len(overloaded)} resources overloaded", 'Yellow', 'Capacity planning needed'])
            
        # Write alerts
        alert_headers = ['Type', 'Issue', 'Severity', 'Action Required']
        ws.append([])  # Empty row
        for col_idx, header in enumerate(alert_headers, 1):
            cell = ws.cell(row=12, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        for idx, alert in enumerate(alerts, 13):
            for col_idx, value in enumerate(alert, 1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                if col_idx == 3:  # Severity column
                    if value == 'Red':
                        cell.fill = self.fills['danger']
                    elif value == 'Orange':
                        cell.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
                    elif value == 'Yellow':
                        cell.fill = self.fills['warning']
                        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
    def get_health_status(self, score):
        """Get health status based on score"""
        if score >= 80:
            return 'Green'
        elif score >= 60:
            return 'Yellow'
        else:
            return 'Red'
            
    def create_impact_matrix_sheet(self):
        """Create 2x2 impact vs risk matrix"""
        ws = self.wb.create_sheet("Impact Matrix")
        
        ws['A1'] = 'IMPACT vs RISK MATRIX'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # Categorize tasks into quadrants
        quadrants = {
            'Critical Focus': [],    # High Impact, High Risk
            'Quick Wins': [],        # High Impact, Low Risk
            'Risk Mitigation': [],   # Low Impact, High Risk
            'Routine': []           # Low Impact, Low Risk
        }
        
        # Threshold for high/low
        impact_threshold = 60
        risk_threshold = 40
        
        for task in self.data:
            impact = task.get('Impact Score', 0)
            risk = task.get('Risk Score', 0)
            
            if impact >= impact_threshold and risk >= risk_threshold:
                quadrants['Critical Focus'].append(task)
            elif impact >= impact_threshold and risk < risk_threshold:
                quadrants['Quick Wins'].append(task)
            elif impact < impact_threshold and risk >= risk_threshold:
                quadrants['Risk Mitigation'].append(task)
            else:
                quadrants['Routine'].append(task)
                
        # Write quadrant summaries
        row = 3
        for quadrant, tasks in quadrants.items():
            ws.cell(row=row, column=1, value=quadrant).font = self.fonts['header']
            ws.cell(row=row, column=2, value=f"({len(tasks)} tasks)")
            ws.merge_cells(f'A{row}:B{row}')
            
            # Color code quadrants
            if quadrant == 'Critical Focus':
                ws.cell(row=row, column=1).fill = self.fills['danger']
            elif quadrant == 'Quick Wins':
                ws.cell(row=row, column=1).fill = self.fills['success']
            elif quadrant == 'Risk Mitigation':
                ws.cell(row=row, column=1).fill = self.fills['warning']
            
            row += 1
            
            # Headers for task details
            headers = ['Task ID', 'Task Name', 'Impact', 'Risk', 'Priority', 'Owner', 'Status']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.fonts['body_bold']
                cell.fill = self.fills['header']
            row += 1
            
            # List top 5 tasks in each quadrant
            for task in sorted(tasks, key=lambda x: x.get('Priority Score', 0), reverse=True)[:5]:
                task_data = [
                    task.get('Task ID', ''),
                    task.get('Task Name', '')[:30],
                    task.get('Impact Score', 0),
                    task.get('Risk Score', 0),
                    task.get('Priority Score', 0),
                    task.get('Resource Assignment', '')[:15],
                    task.get('Status', '')
                ]
                for col_idx, value in enumerate(task_data, 1):
                    ws.cell(row=row, column=col_idx, value=value)
                row += 1
                
            row += 2  # Space between quadrants
            
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        
    def create_resource_orchestra_sheet(self):
        """Create resource utilization and capacity view"""
        ws = self.wb.create_sheet("Resource Orchestra")
        
        ws['A1'] = 'RESOURCE CAPACITY & ALLOCATION'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:G1')
        
        # Aggregate resource data
        resource_data = {}
        for task in self.data:
            resource = task.get('Resource Assignment', 'Unassigned')
            if resource not in resource_data:
                resource_data[resource] = {
                    'tasks': 0,
                    'load': [],
                    'critical_tasks': 0,
                    'blocked_tasks': 0,
                    'total_budget': 0,
                    'spent': 0
                }
            
            resource_data[resource]['tasks'] += 1
            resource_data[resource]['load'].append(task.get('Resource Load %', 0))
            if task.get('Criticality Level') == 'Critical':
                resource_data[resource]['critical_tasks'] += 1
            if task.get('Status') == 'Blocked':
                resource_data[resource]['blocked_tasks'] += 1
            resource_data[resource]['total_budget'] += task.get('Cost Budget', 0)
            resource_data[resource]['spent'] += task.get('Actual Cost', 0)
            
        # Write resource summary
        row = 3
        headers = ['Resource/Team', 'Active Tasks', 'Avg Load %', 'Critical Tasks', 
                   'Blocked', 'Budget Allocated', 'Budget Used', 'Health']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            
        row = 4
        for resource, data in sorted(resource_data.items(), 
                                    key=lambda x: np.mean(x[1]['load']) if x[1]['load'] else 0, 
                                    reverse=True):
            avg_load = np.mean(data['load']) if data['load'] else 0
            budget_usage = (data['spent'] / data['total_budget'] * 100) if data['total_budget'] > 0 else 0
            
            # Determine health
            if avg_load > 85 or data['blocked_tasks'] > 0:
                health = 'Red'
            elif avg_load > 70 or data['critical_tasks'] > 2:
                health = 'Yellow'
            else:
                health = 'Green'
                
            row_data = [
                resource,
                data['tasks'],
                f"{avg_load:.0f}%",
                data['critical_tasks'],
                data['blocked_tasks'],
                f"${data['total_budget']:,.0f}",
                f"${data['spent']:,.0f} ({budget_usage:.0f}%)",
                health
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.borders['thin']
                
                # Color health column
                if col_idx == 8:
                    if health == 'Green':
                        cell.fill = self.fills['success']
                    elif health == 'Yellow':
                        cell.fill = self.fills['warning']
                    elif health == 'Red':
                        cell.fill = self.fills['danger']
                        
            row += 1
            
        # Bottleneck Analysis
        ws.cell(row=row+2, column=1, value='BOTTLENECK ANALYSIS').font = self.fonts['header']
        row += 4
        
        # Find bottlenecks
        bottlenecks = []
        for resource, data in resource_data.items():
            avg_load = np.mean(data['load']) if data['load'] else 0
            if avg_load > 80 or data['blocked_tasks'] > 0:
                bottlenecks.append({
                    'resource': resource,
                    'issue': 'Overloaded' if avg_load > 80 else 'Has Blocked Tasks',
                    'impact': f"{data['critical_tasks']} critical tasks affected",
                    'recommendation': 'Redistribute load' if avg_load > 80 else 'Remove blockers'
                })
                
        bottle_headers = ['Resource', 'Issue', 'Impact', 'Recommendation']
        for col_idx, header in enumerate(bottle_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        row += 1
        for bottleneck in bottlenecks[:5]:  # Top 5 bottlenecks
            for col_idx, key in enumerate(['resource', 'issue', 'impact', 'recommendation'], 1):
                ws.cell(row=row, column=col_idx, value=bottleneck[key])
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 10
        
    def create_timeline_rhythm_sheet(self):
        """Create timeline and milestone tracking"""
        ws = self.wb.create_sheet("Timeline Rhythm")
        
        ws['A1'] = 'TIMELINE & CRITICAL PATH'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # Get tasks with dates
        timeline_tasks = [t for t in self.data if t.get('Start Date') and t.get('Task Type') != 'Child']
        timeline_tasks.sort(key=lambda x: pd.to_datetime(x['Start Date']))
        
        # Milestone tracker
        ws['A3'] = 'UPCOMING MILESTONES'
        ws['A3'].font = self.fonts['header']
        
        milestones = [t for t in self.data if t.get('Milestone') == 'Yes' and t.get('% Complete', 0) < 100]
        
        row = 5
        mile_headers = ['Milestone', 'Target Date', 'Days Until', 'Dependencies', 'Status', 'Impact if Delayed']
        for col_idx, header in enumerate(mile_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        row = 6
        for milestone in milestones[:5]:
            try:
                target_date = pd.to_datetime(milestone.get('End Date', milestone.get('Start Date')))
                if pd.isna(target_date):
                    target_date = pd.to_datetime(milestone.get('Start Date', '2024-01-01'))
                days_until = (target_date - pd.Timestamp.now()).days
                date_str = target_date.strftime('%Y-%m-%d')
            except:
                target_date = pd.Timestamp.now()
                days_until = 0
                date_str = 'TBD'
            
            milestone_data = [
                milestone.get('Task Name', ''),
                date_str,
                days_until,
                len(milestone.get('Dependencies', '').split(',')) if milestone.get('Dependencies') else 0,
                milestone.get('Status', ''),
                'High - Blocks multiple tasks' if milestone.get('Blocking Tasks') else 'Medium'
            ]
            
            for col_idx, value in enumerate(milestone_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                # Color code days until
                if col_idx == 3:
                    if value < 0:
                        cell.fill = self.fills['danger']
                    elif value < 7:
                        cell.fill = self.fills['warning']
                    else:
                        cell.fill = self.fills['success']
            row += 1
            
        # Critical Path Tasks
        ws.cell(row=row+2, column=1, value='CRITICAL PATH TASKS').font = self.fonts['header']
        row += 4
        
        cp_tasks = [t for t in self.data if t.get('Critical Path') == 'Yes']
        cp_headers = ['Task ID', 'Task Name', 'Start', 'End', 'Progress', 'Float', 'Status']
        
        for col_idx, header in enumerate(cp_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        row += 1
        for task in cp_tasks[:10]:
            task_data = [
                task.get('Task ID', ''),
                task.get('Task Name', '')[:40],
                task.get('Start Date', ''),
                task.get('End Date', ''),
                f"{task.get('% Complete', 0)}%",
                task.get('Total Float', 0),
                task.get('Status', '')
            ]
            
            for col_idx, value in enumerate(task_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                # Highlight delays
                if col_idx == 7 and value == 'Delayed':
                    cell.fill = self.fills['danger']
            row += 1
            
        # Velocity Metrics
        ws.cell(row=row+2, column=1, value='PROJECT VELOCITY').font = self.fonts['header']
        row += 4
        
        velocity_data = [
            ['Planned Completion', len([t for t in self.data if t.get('Status') == 'Complete'])],
            ['Actual Completion', len([t for t in self.data if t.get('Status') == 'Complete'])],
            ['Tasks Behind Schedule', len([t for t in self.data if t.get('Variance Days', 0) < 0])],
            ['Average Delay (days)', np.mean([abs(t.get('Variance Days', 0)) for t in self.data if t.get('Variance Days', 0) < 0]) if any(t.get('Variance Days', 0) < 0 for t in self.data) else 0],
            ['Projected Completion', 'On Time' if np.mean([t.get('SPI', 1) for t in self.data]) >= 0.95 else 'Delayed']
        ]
        
        for vel_data in velocity_data:
            for col_idx, value in enumerate(vel_data, 1):
                ws.cell(row=row, column=col_idx, value=value)
            row += 1
            
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
            
    def create_decision_command_sheet(self):
        """Create decision and action tracking"""
        ws = self.wb.create_sheet("Decision Command")
        
        ws['A1'] = 'DECISIONS & ACTIONS REQUIRED'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:F1')
        
        # Collect decisions needed
        decisions = []
        
        # Blocked tasks need decisions
        for task in self.data:
            if task.get('Status') == 'Blocked':
                decisions.append({
                    'type': 'UNBLOCK',
                    'urgency': 'Critical',
                    'task': task.get('Task Name', ''),
                    'action': f"Remove blocker for {task.get('Task ID', '')}",
                    'owner': task.get('Resource Assignment', ''),
                    'impact': 'Project delay if not resolved',
                    'due': 'Immediate'
                })
                
            if task.get('Risk Score', 0) > 70:
                decisions.append({
                    'type': 'MITIGATE',
                    'urgency': 'High',
                    'task': task.get('Task Name', ''),
                    'action': f"Implement risk mitigation for {task.get('Task ID', '')}",
                    'owner': task.get('Resource Assignment', ''),
                    'impact': task.get('Risk Mitigation', 'Potential project impact'),
                    'due': 'This week'
                })
                
            if task.get('Resource Load %', 0) > 90:
                decisions.append({
                    'type': 'RESOURCE',
                    'urgency': 'Medium',
                    'task': task.get('Task Name', ''),
                    'action': 'Reallocate resources',
                    'owner': 'Project Manager',
                    'impact': 'Resource burnout risk',
                    'due': 'Next sprint'
                })
                
        # Write decision queue
        row = 3
        headers = ['Type', 'Urgency', 'Task', 'Action Required', 'Owner', 'Impact if Delayed', 'Due']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            
        row = 4
        # Sort by urgency
        urgency_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
        decisions.sort(key=lambda x: urgency_order.get(x['urgency'], 4))
        
        for decision in decisions[:10]:  # Top 10 decisions
            for col_idx, key in enumerate(['type', 'urgency', 'task', 'action', 'owner', 'impact', 'due'], 1):
                cell = ws.cell(row=row, column=col_idx, value=decision.get(key, ''))
                cell.border = self.borders['thin']
                
                # Color code urgency
                if col_idx == 2:
                    if decision['urgency'] == 'Critical':
                        cell.fill = self.fills['danger']
                    elif decision['urgency'] == 'High':
                        cell.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
                    elif decision['urgency'] == 'Medium':
                        cell.fill = self.fills['warning']
                        
            row += 1
            
        # Quick Actions Summary
        ws.cell(row=row+2, column=1, value='QUICK ACTIONS SUMMARY').font = self.fonts['header']
        row += 4
        
        action_summary = [
            ['Total Decisions Pending', len(decisions)],
            ['Critical Actions', len([d for d in decisions if d['urgency'] == 'Critical'])],
            ['Blockers to Remove', len([d for d in decisions if d['type'] == 'UNBLOCK'])],
            ['Risk Mitigations', len([d for d in decisions if d['type'] == 'MITIGATE'])],
            ['Resource Issues', len([d for d in decisions if d['type'] == 'RESOURCE'])]
        ]
        
        for summary in action_summary:
            for col_idx, value in enumerate(summary, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if col_idx == 1:
                    cell.font = self.fonts['body_bold']
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 12
        
    def create_predictive_insights_sheet(self):
        """Create AI-powered predictions and recommendations"""
        ws = self.wb.create_sheet("Predictive Insights")
        
        ws['A1'] = 'PREDICTIVE ANALYTICS & RECOMMENDATIONS'
        ws['A1'].font = Font(name='Verdana', size=16, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:F1')
        
        # Calculate predictions
        avg_spi = np.mean([t.get('SPI', 1) for t in self.data])
        avg_cpi = np.mean([t.get('CPI', 1) for t in self.data])
        completion_rate = len([t for t in self.data if t['Status'] == 'Complete']) / len(self.data)
        risk_trend = np.mean([t.get('Risk Score', 0) for t in self.data])
        
        # Completion confidence
        confidence = min(100, max(0, 
            (avg_spi * 30) + 
            (avg_cpi * 30) + 
            (completion_rate * 20) + 
            ((100 - risk_trend) / 100 * 20)
        ))
        
        # Predictions
        ws['A3'] = 'COMPLETION PREDICTIONS'
        ws['A3'].font = self.fonts['header']
        
        predictions = [
            ['Metric', 'Current', 'Predicted', 'Confidence', 'Trend'],
            ['Project Completion Date', 'Apr 30, 2024', 'May 15, 2024' if avg_spi < 0.95 else 'Apr 30, 2024', f"{confidence:.0f}%", '↓' if avg_spi < 0.95 else '→'],
            ['Final Budget', '$8.5M', f'${8.5 * (1/avg_cpi):.1f}M' if avg_cpi < 1 else '$8.5M', f"{min(100, avg_cpi*100):.0f}%", '↑' if avg_cpi < 1 else '→'],
            ['Tasks at Risk', len([t for t in self.data if t.get('Risk Score', 0) > 40]), int(len(self.data) * 0.25), '75%', '↑'],
            ['Resource Capacity', '78%', '85%', '80%', '↑']
        ]
        
        row = 5
        for pred_row in predictions:
            for col_idx, value in enumerate(pred_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if row == 5:  # Header
                    cell.font = self.fonts['body_bold']
                    cell.fill = self.fills['header']
                cell.border = self.borders['thin']
            row += 1
            
        # Risk Predictions
        ws.cell(row=row+2, column=1, value='EMERGING RISKS').font = self.fonts['header']
        row += 4
        
        risk_predictions = [
            {
                'risk': 'Security Framework Delay',
                'probability': '65%',
                'impact': 'High',
                'timeline': 'Next 2 weeks',
                'mitigation': 'Add security resources immediately'
            },
            {
                'risk': 'Resource Burnout',
                'probability': '45%',
                'impact': 'Medium',
                'timeline': 'Next month',
                'mitigation': 'Implement resource rotation plan'
            },
            {
                'risk': 'Budget Overrun',
                'probability': '30%',
                'impact': 'Medium',
                'timeline': 'Q2 2024',
                'mitigation': 'Review and optimize spending'
            }
        ]
        
        risk_headers = ['Risk', 'Probability', 'Impact', 'Timeline', 'Recommended Action']
        for col_idx, header in enumerate(risk_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        row += 1
        for risk in risk_predictions:
            for col_idx, key in enumerate(['risk', 'probability', 'impact', 'timeline', 'mitigation'], 1):
                cell = ws.cell(row=row, column=col_idx, value=risk[key])
                # Color impact
                if col_idx == 3:
                    if risk['impact'] == 'High':
                        cell.fill = self.fills['danger']
                    elif risk['impact'] == 'Medium':
                        cell.fill = self.fills['warning']
            row += 1
            
        # Optimization Recommendations
        ws.cell(row=row+2, column=1, value='OPTIMIZATION OPPORTUNITIES').font = self.fonts['header']
        row += 4
        
        optimizations = [
            ['Resource Reallocation', 'Move 2 developers from Phase 2 to Phase 3', '15% faster delivery', 'High'],
            ['Parallel Execution', 'Run testing in parallel with development', '10 days saved', 'Medium'],
            ['Scope Adjustment', 'Defer 2 non-critical features to Phase 2', '20% risk reduction', 'Medium'],
            ['Tool Automation', 'Implement automated testing for APIs', '30% effort reduction', 'High']
        ]
        
        opt_headers = ['Opportunity', 'Action', 'Expected Benefit', 'Confidence']
        for col_idx, header in enumerate(opt_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts['body_bold']
            cell.fill = self.fills['header']
            
        row += 1
        for opt in optimizations:
            for col_idx, value in enumerate(opt, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if col_idx == 4 and value == 'High':
                    cell.fill = self.fills['success']
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 35
        
    def save_template(self, filename='working_project_template.xlsx'):
        """Save the template"""
        self.create_main_sheet()
        self.create_dashboard_sheet()
        self.create_instructions_sheet()
        
        # Add sync-up dashboard sheets
        self.create_syncup_dashboard_sheets()
        
        # Save workbook
        self.wb.save(filename)
        print(f"Working template saved as: {filename}")
        
        # Also save as CSV for easy viewing
        df = pd.DataFrame(self.data)
        csv_filename = filename.replace('.xlsx', '.csv')
        df.to_csv(csv_filename, index=False)
        print(f"Data also saved as: {csv_filename}")
        

if __name__ == "__main__":
    generator = WorkingProjectTemplate()
    generator.save_template('/Users/haithamdata/Documents/Prog/My Productivity/Project Managment/Project managment tamblet/syncup_dashboard_template.xlsx')
    generator.save_template('/Users/haithamdata/Documents/Prog/My Productivity/Project Managment/Project managment tamblet/working_project_template.xlsx')