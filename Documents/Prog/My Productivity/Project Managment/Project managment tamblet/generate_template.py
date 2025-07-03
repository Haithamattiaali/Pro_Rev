#!/usr/bin/env python3
"""
Project Management Template Generator
Creates an Excel template with impact-based progress tracking
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime, timedelta
import numpy as np

class ProjectTemplateGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Define color scheme and styles"""
        self.colors = {
            'primary': '9E1F63',
            'secondary': '721548',
            'accent_blue': '005B8C',
            'accent_coral': 'E05E3D',
            'status_green': '2ECC71',
            'status_yellow': 'F1C40F',
            'status_red': 'E74C3C',
            'light_gray': 'F8F9FA',
            'medium_gray': 'E9ECEF'
        }
        
        self.fonts = {
            'header': Font(name='Verdana', size=14, bold=True, color=self.colors['primary']),
            'subheader': Font(name='Verdana', size=12, bold=True, color=self.colors['secondary']),
            'body': Font(name='Verdana', size=10),
            'body_bold': Font(name='Verdana', size=10, bold=True),
            'impact': Font(name='Verdana', size=11, bold=True)
        }
        
        self.fills = {
            'header': PatternFill(start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid'),
            'parent': PatternFill(start_color=self.colors['medium_gray'], end_color=self.colors['medium_gray'], fill_type='solid'),
            'green': PatternFill(start_color=self.colors['status_green'], end_color=self.colors['status_green'], fill_type='solid'),
            'yellow': PatternFill(start_color=self.colors['status_yellow'], end_color=self.colors['status_yellow'], fill_type='solid'),
            'red': PatternFill(start_color=self.colors['status_red'], end_color=self.colors['status_red'], fill_type='solid')
        }
        
        self.borders = {
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'thick': Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        }
        
    def create_main_sheet(self):
        """Create the main project tracking sheet"""
        ws = self.wb.active
        ws.title = "Project Tasks"
        
        # Define headers
        headers = [
            'Task ID', 'Task Name', 'Task Type', 'Parent Task ID',
            'Duration (Days)', 'Start Date', 'End Date', 'Actual Start',
            'Actual End', '% Complete', 'Task Agility', 'Dependencies',
            'Dependency Type', 'Impact Score', 'Criticality Level',
            'Resource Assignment', 'Status', 'Health Indicator', 'Notes',
            'Blocking Tasks', 'Blocked By', 'Total Float', 'Free Float',
            'Weighted Progress', 'Rolled Up Progress'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.borders['thick']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths
        column_widths = {
            'A': 10, 'B': 30, 'C': 12, 'D': 15, 'E': 12, 'F': 12,
            'G': 12, 'H': 12, 'I': 12, 'J': 10, 'K': 12, 'L': 15,
            'M': 15, 'N': 12, 'O': 15, 'P': 20, 'Q': 15, 'R': 15,
            'S': 25, 'T': 12, 'U': 12, 'V': 10, 'W': 10, 'X': 15, 'Y': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add data validations
        self.add_validations(ws)
        
        # Add sample data
        self.add_sample_data(ws)
        
        # Add formulas
        self.add_formulas(ws)
        
        # Add conditional formatting
        self.add_conditional_formatting(ws)
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
    def add_validations(self, ws):
        """Add dropdown validations"""
        # Task Type validation
        task_type_val = DataValidation(
            type="list",
            formula1='"Parent,Child"',
            allow_blank=False
        )
        task_type_val.add('C2:C1000')
        ws.add_data_validation(task_type_val)
        
        # Task Agility validation
        agility_val = DataValidation(
            type="list",
            formula1='"Parallel,Sequential"',
            allow_blank=False
        )
        agility_val.add('K2:K1000')
        ws.add_data_validation(agility_val)
        
        # Dependency Type validation
        dep_type_val = DataValidation(
            type="list",
            formula1='"FS,SS,FF,SF"',
            allow_blank=True
        )
        dep_type_val.add('M2:M1000')
        ws.add_data_validation(dep_type_val)
        
        # Criticality Level validation
        criticality_val = DataValidation(
            type="list",
            formula1='"High,Medium,Low"',
            allow_blank=False
        )
        criticality_val.add('O2:O1000')
        ws.add_data_validation(criticality_val)
        
        # Status validation
        status_val = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Delayed,Blocked"',
            allow_blank=False
        )
        status_val.add('Q2:Q1000')
        ws.add_data_validation(status_val)
        
        # Health Indicator validation
        health_val = DataValidation(
            type="list",
            formula1='"Green,Yellow,Red"',
            allow_blank=False
        )
        health_val.add('R2:R1000')
        ws.add_data_validation(health_val)
        
    def add_sample_data(self, ws):
        """Add sample project data"""
        sample_data = [
            # Parent tasks
            ['P001', 'Project Initiation', 'Parent', '', 10, '2024-01-01', '2024-01-10', '', '', 0, 'Sequential', '', '', '', 'High', 'PM Team', 'Not Started', 'Green', 'Project kickoff phase'],
            ['P002', 'Requirements Gathering', 'Parent', '', 15, '2024-01-11', '2024-01-25', '', '', 0, 'Sequential', 'P001', 'FS', '', 'High', 'Business Analysts', 'Not Started', 'Green', 'Gather all requirements'],
            ['P003', 'Design Phase', 'Parent', '', 20, '2024-01-26', '2024-02-14', '', '', 0, 'Parallel', 'P002', 'FS', '', 'High', 'Design Team', 'Not Started', 'Green', 'System design'],
            ['P004', 'Development', 'Parent', '', 40, '2024-02-15', '2024-03-25', '', '', 0, 'Parallel', 'P003', 'FS', '', 'High', 'Dev Team', 'Not Started', 'Green', 'Build the system'],
            ['P005', 'Testing', 'Parent', '', 15, '2024-03-26', '2024-04-09', '', '', 0, 'Sequential', 'P004', 'FS', '', 'High', 'QA Team', 'Not Started', 'Green', 'System testing'],
            
            # Child tasks for P001
            ['C001', 'Form Project Team', 'Child', 'P001', 3, '2024-01-01', '2024-01-03', '', '', 0, 'Sequential', '', '', '', 'High', 'HR', 'Not Started', 'Green', 'Assemble team'],
            ['C002', 'Define Project Charter', 'Child', 'P001', 2, '2024-01-04', '2024-01-05', '', '', 0, 'Sequential', 'C001', 'FS', '', 'High', 'PM', 'Not Started', 'Green', 'Create charter'],
            ['C003', 'Stakeholder Analysis', 'Child', 'P001', 2, '2024-01-06', '2024-01-07', '', '', 0, 'Parallel', 'C001', 'FS', '', 'Medium', 'PM', 'Not Started', 'Green', 'Identify stakeholders'],
            ['C004', 'Kickoff Meeting', 'Child', 'P001', 1, '2024-01-08', '2024-01-08', '', '', 0, 'Sequential', 'C002,C003', 'FS', '', 'High', 'All', 'Not Started', 'Green', 'Project kickoff'],
            
            # Child tasks for P002
            ['C005', 'Business Requirements', 'Child', 'P002', 5, '2024-01-11', '2024-01-15', '', '', 0, 'Sequential', 'P001', 'FS', '', 'High', 'BA', 'Not Started', 'Green', 'Gather business reqs'],
            ['C006', 'Technical Requirements', 'Child', 'P002', 5, '2024-01-16', '2024-01-20', '', '', 0, 'Sequential', 'C005', 'FS', '', 'High', 'Tech Lead', 'Not Started', 'Green', 'Define tech reqs'],
            ['C007', 'Requirements Review', 'Child', 'P002', 3, '2024-01-21', '2024-01-23', '', '', 0, 'Sequential', 'C005,C006', 'FS', '', 'High', 'All', 'Not Started', 'Green', 'Review all reqs'],
            ['C008', 'Requirements Sign-off', 'Child', 'P002', 2, '2024-01-24', '2024-01-25', '', '', 0, 'Sequential', 'C007', 'FS', '', 'High', 'Stakeholders', 'Not Started', 'Green', 'Get approval'],
        ]
        
        # Write sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.borders['thin']
                
                # Apply special formatting for parent tasks
                if row_data[2] == 'Parent':
                    cell.fill = self.fills['parent']
                    cell.font = self.fonts['body_bold']
                else:
                    cell.font = self.fonts['body']
                    
    def add_formulas(self, ws):
        """Add Excel formulas for calculations"""
        # Starting from row 2 (after headers)
        for row in range(2, 20):  # Adjust range as needed
            # End Date formula (Start Date + Duration)
            ws[f'G{row}'] = f'=IF(AND(F{row}<>"",E{row}<>""),WORKDAY(F{row},E{row}-1),"")'
            
            # % Complete formula (based on actual dates)
            ws[f'J{row}'] = f'=IF(Q{row}="Complete",100,IF(AND(H{row}<>"",H{row}<=TODAY()),MIN((TODAY()-H{row})/(G{row}-F{row})*100,99),0))'
            
            # Impact Score formula
            impact_formula = self.create_impact_formula(row)
            ws[f'N{row}'] = impact_formula
            
            # Health Indicator formula
            health_formula = f'''=IF(Q{row}="Complete","Green",
                IF(OR(Q{row}="Delayed",Q{row}="Blocked"),"Red",
                IF(AND(J{row}<100,G{row}<TODAY()),"Red",
                IF(AND(J{row}<80,G{row}<=TODAY()+7),"Yellow","Green"))))'''
            ws[f'R{row}'] = health_formula
            
            # Weighted Progress (for parent tasks)
            ws[f'X{row}'] = f'=IF(C{row}="Parent",SUMPRODUCT((D2:D100=A{row})*(J2:J100)*(N2:N100))/SUMIF(D2:D100,A{row},N2:N100),J{row})'
            
            # Rolled Up Progress
            ws[f'Y{row}'] = f'=IF(C{row}="Child",J{row},X{row})'
            
    def create_impact_formula(self, row):
        """Create the impact score formula"""
        formula = f'''=ROUND(
            (
                (E{row}/MAX(E:E)*25) +
                (LEN(L{row})-LEN(SUBSTITUTE(L{row},",",""))+1)*30/10 +
                (IF(O{row}="High",100,IF(O{row}="Medium",60,30))*0.2) +
                ((1-ROW()/COUNTA(A:A))*25)
            ) *
            IF(LEN(T{row})>0,1.5,1) *
            IF(K{row}="Parallel",0.7,1) *
            IF(Q{row}="Delayed",1.3,IF(Q{row}="Blocked",1.5,1))
        ,0)'''
        return formula
        
    def add_conditional_formatting(self, ws):
        """Add conditional formatting rules"""
        # Health indicator colors
        ws.conditional_formatting.add('R2:R1000',
            CellIsRule(operator='equal', formula=['"Green"'], fill=self.fills['green']))
        ws.conditional_formatting.add('R2:R1000',
            CellIsRule(operator='equal', formula=['"Yellow"'], fill=self.fills['yellow']))
        ws.conditional_formatting.add('R2:R1000',
            CellIsRule(operator='equal', formula=['"Red"'], fill=self.fills['red']))
        
        # Progress color scale
        ws.conditional_formatting.add('J2:J1000',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',
                mid_type='num', mid_value=50, mid_color='FFFF00',
                end_type='num', end_value=100, end_color='00FF00'
            ))
        
        # Impact score color scale
        ws.conditional_formatting.add('N2:N1000',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='00FF00',
                mid_type='num', mid_value=50, mid_color='FFFF00',
                end_type='num', end_value=100, end_color='FF0000'
            ))
            
    def create_dashboard_sheet(self):
        """Create executive dashboard sheet"""
        ws = self.wb.create_sheet("Dashboard")
        
        # Title
        ws['A1'] = 'PROJECT EXECUTIVE DASHBOARD'
        ws['A1'].font = Font(name='Verdana', size=20, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:H1')
        
        # Key Metrics Section
        ws['A3'] = 'KEY METRICS'
        ws['A3'].font = self.fonts['header']
        
        metrics = [
            ['Overall Progress:', '=AVERAGE(\'Project Tasks\'!Y:Y)'],
            ['Critical Tasks:', '=COUNTIF(\'Project Tasks\'!O:O,"High")'],
            ['Delayed Tasks:', '=COUNTIF(\'Project Tasks\'!Q:Q,"Delayed")'],
            ['Blocked Tasks:', '=COUNTIF(\'Project Tasks\'!Q:Q,"Blocked")'],
            ['Average Impact Score:', '=AVERAGE(\'Project Tasks\'!N:N)'],
            ['Health Status:', '=COUNTIF(\'Project Tasks\'!R:R,"Red")&" Red, "&COUNTIF(\'Project Tasks\'!R:R,"Yellow")&" Yellow, "&COUNTIF(\'Project Tasks\'!R:R,"Green")&" Green"']
        ]
        
        for idx, (label, formula) in enumerate(metrics, 5):
            ws[f'A{idx}'] = label
            ws[f'C{idx}'] = formula
            ws[f'A{idx}'].font = self.fonts['body_bold']
            
        # Critical Path Section
        ws['A12'] = 'TOP 5 CRITICAL TASKS BY IMPACT'
        ws['A12'].font = self.fonts['header']
        
        # Headers for critical tasks
        crit_headers = ['Task Name', 'Impact Score', 'Status', 'Health']
        for col, header in enumerate(crit_headers, 1):
            ws.cell(row=14, column=col, value=header).font = self.fonts['subheader']
            
    def create_instructions_sheet(self):
        """Create instructions sheet"""
        ws = self.wb.create_sheet("Instructions")
        
        instructions = [
            ['PROJECT MANAGEMENT TEMPLATE - USER GUIDE', ''],
            ['', ''],
            ['OVERVIEW:', 'This template uses an impact-based scoring system to provide accurate project progress tracking.'],
            ['', ''],
            ['KEY FEATURES:', ''],
            ['1. Impact Score Algorithm:', 'Calculates task importance based on duration, dependencies, criticality, and timeline position'],
            ['2. Parent-Child Hierarchy:', 'Progress rolls up from child tasks to parent tasks weighted by impact scores'],
            ['3. Automatic Health Tracking:', 'Health indicators update based on progress and deadlines'],
            ['', ''],
            ['HOW TO USE:', ''],
            ['1. Enter Tasks:', 'Start by entering parent tasks, then add child tasks with parent IDs'],
            ['2. Set Dependencies:', 'Link tasks using Task IDs in the Dependencies column'],
            ['3. Assign Resources:', 'Add team members or departments to Resource Assignment'],
            ['4. Track Progress:', 'Update Actual Start dates and Status - progress calculates automatically'],
            ['5. Monitor Health:', 'Watch the Health Indicator column for risk identification'],
            ['', ''],
            ['IMPACT SCORE COMPONENTS:', ''],
            ['- Duration Weight (25%):', 'Longer tasks have higher impact'],
            ['- Dependency Count (30%):', 'More dependencies increase impact'],
            ['- Resource Criticality (20%):', 'High criticality resources increase impact'],
            ['- Timeline Position (25%):', 'Earlier tasks have higher impact'],
            ['', ''],
            ['BEST PRACTICES:', ''],
            ['1. Update Daily:', 'Keep actual dates and status current for accurate tracking'],
            ['2. Review Impact Scores:', 'Focus on high-impact tasks for maximum efficiency'],
            ['3. Use Filters:', 'Filter by Status, Health, or Resource for focused views'],
            ['4. Check Dashboard:', 'Review executive dashboard for quick project health assessment']
        ]
        
        for row_idx, (label, desc) in enumerate(instructions, 1):
            ws[f'A{row_idx}'] = label
            ws[f'B{row_idx}'] = desc
            if label and not desc:  # Headers
                ws[f'A{row_idx}'].font = self.fonts['header']
            elif label.endswith(':'):  # Section headers
                ws[f'A{row_idx}'].font = self.fonts['subheader']
            else:
                ws[f'A{row_idx}'].font = self.fonts['body_bold']
                ws[f'B{row_idx}'].font = self.fonts['body']
                
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        
    def save_template(self, filename='project_management_template.xlsx'):
        """Save the template"""
        self.create_main_sheet()
        self.create_dashboard_sheet()
        self.create_instructions_sheet()
        
        # Save the workbook
        self.wb.save(filename)
        print(f"Template saved as: {filename}")
        

if __name__ == "__main__":
    generator = ProjectTemplateGenerator()
    generator.save_template('/Users/haithamdata/Documents/Prog/My Productivity/Project Managment/Project managment tamblet/project_management_template.xlsx')